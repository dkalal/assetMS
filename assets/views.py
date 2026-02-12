from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.views.decorators.http import require_http_methods, require_GET
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError, PermissionDenied
from django.views.decorators.http import require_GET, require_POST
from django.db.models import Count, Q
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie, csrf_protect
from django.utils.decorators import method_decorator
import re
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.urls import reverse_lazy
from django.views.generic import CreateView, ListView, DetailView, TemplateView, UpdateView
from .models import Asset, AssetCategory, AssetCategoryField, ExportLog
from .forms import AssetForm

from django.urls import NoReverseMatch
import qrcode
from io import BytesIO
from django.core.files.base import ContentFile
from django.http import JsonResponse
from django.views.decorators.http import require_GET, require_POST
from django.db.models import Count, Q
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie, csrf_protect
from django.utils.decorators import method_decorator
import re
from django.http import HttpResponse
import csv
import pandas as pd
from django.conf import settings
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
from datetime import datetime, date
import openpyxl
from openpyxl.utils.datetime import from_excel
from django.core.files.storage import default_storage
from django.contrib import messages
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from audit.models import AuditLog
from audit.utils import log_audit, ASSIGN_ACTION, MAINTENANCE_ACTION
from django.core.paginator import Paginator, EmptyPage
from django.utils.timezone import localtime

from users.utils import can
from tenancy.mixins import BranchContextMixin, CompanyScopedQuerysetMixin, company_required

# Permission check: only admin/manager
def is_admin_or_manager(user):
    return user.is_authenticated and user.role in ('admin', 'manager')

# Create your views here.

class AssetCreateView(LoginRequiredMixin, BranchContextMixin, TemplateView):
    """
    Unified Asset Creation View with Smart Routing.
    
    World-Class Implementation:
    - Single entry point for all asset creation
    - Automatic role-based workflow routing:
      * Admin → Direct creation (no approval)
      * Manager → Approval workflow (if company policy requires)
      * User → Permission denied
    - Single form, single URL, smart backend
    - Consistent UX across all roles
    
    URL: /assets/register/
    Template: assets/asset_form.html
    """
    template_name = 'assets/asset_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        """Check permissions and determine workflow mode."""
        user = request.user
        
        # Only admins and managers can create assets
        if user.role not in ['admin', 'manager']:
            messages.error(request, "You don't have permission to create assets.")
            return redirect('asset_list')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        """Prepare context with workflow mode and form."""
        context = super().get_context_data(**kwargs)
        user = self.request.user
        company = getattr(self.request, 'company', None)
        
        # Determine workflow mode
        if user.role == 'admin':
            # Admins always create directly
            context['workflow_mode'] = 'direct'
            context['requires_approval'] = False
            context['workflow_message'] = 'Asset will be created immediately'
        elif user.role == 'manager':
            # Check company policy for managers
            # Default: require approval for managers
            requires_approval = True  # Can be made configurable per company
            context['workflow_mode'] = 'approval' if requires_approval else 'direct'
            context['requires_approval'] = requires_approval
            if requires_approval:
                context['workflow_message'] = 'This request will be sent to an admin for approval'
            else:
                context['workflow_message'] = 'Asset will be created immediately'
        
        context['user_role'] = user.role
        
        # Provide form instance for template
        if 'form' not in context:
            from .forms import AssetForm
            context['form'] = AssetForm(request=self.request)
        
        return context
    
    def get(self, request, *args, **kwargs):
        """Display the unified creation form."""
        return self.render_to_response(self.get_context_data())
    
    def post(self, request, *args, **kwargs):
        """Handle form submission with smart routing."""
        user = request.user
        company = getattr(request, 'company', None)
        
        # Determine workflow
        if user.role == 'admin':
            # Admin: Direct creation
            return self.create_asset_directly(request)
        elif user.role == 'manager':
            # Manager: Check if approval required
            requires_approval = True  # Can be made configurable
            if requires_approval:
                return self.create_approval_request(request)
            else:
                return self.create_asset_directly(request)
        else:
            return JsonResponse({
                'success': False,
                'error': 'Permission denied'
            }, status=403)
    
    def create_asset_directly(self, request):
        """Create asset directly without approval."""
        from .forms import AssetForm
        
        form = AssetForm(request.POST, request.FILES, request=request)
        
        if not form.is_valid():
            messages.error(request, "Please correct the errors below.")
            context = self.get_context_data()
            context['form'] = form
            return self.render_to_response(context)
        
        # Save asset
        asset = form.save(commit=False)
        company = asset.company
        branch = asset.branch
        
        if asset.company and not asset.company_id:
            asset.company_id = asset.company.id
        
        assigned_to = form.cleaned_data.get('assigned_to')
        asset.save()
        
        # WORLD-CLASS QR CODE GENERATION with duplicate collision detection
        try:
            import os
            from django.conf import settings
            
            qr_dir = os.path.join(settings.MEDIA_ROOT, 'qr_codes')
            os.makedirs(qr_dir, exist_ok=True)
            
            base_url = request.build_absolute_uri('/')[:-1]
            qr_url = f"{base_url}/assets/{asset.uuid}/"
            
            # Store QR string for duplicate detection
            asset.qr_string = qr_url
            
            # Generate QR code image
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_H,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_url)
            qr.make(fit=True)
            
            qr_img = qr.make_image(fill_color="black", back_color="white")
            buffer = BytesIO()
            qr_img.save(buffer, 'PNG')
            asset.qr_code.save(f"asset_{asset.uuid}.png", ContentFile(buffer.getvalue()), save=False)
        except Exception as e:
            print(f"QR code generation failed: {e}")
            # Continue without QR code - not critical for asset creation
        
        asset.save()
        
        # Audit logs
        if assigned_to:
            log_audit(
                request.user,
                ASSIGN_ACTION,
                asset,
                f'Asset assigned to {assigned_to}',
                related_user=assigned_to,
                company=company,
                branch=branch
            )
        
        log_audit(
            request.user,
            'create',
            asset,
            f'Asset created in branch: {branch.name}',
            company=company,
            branch=branch
        )
        
        messages.success(request, f"Asset '{asset}' registered successfully in {branch.name}.")
        return redirect('asset_list')
    
    def create_approval_request(self, request):
        """Create approval request for manager workflow."""
        from tenancy.approval_models import ApprovalRequest
        from tenancy.models import Alert, Branch
        from .models import AssetCategory, AssetCategoryField
        
        try:
            with transaction.atomic():
                company = getattr(request, 'company', None)
                user = request.user
                
                # Extract form data
                title = request.POST.get('title', '').strip()
                description = request.POST.get('description', '').strip()
                category_id = request.POST.get('category')
                branch_id = request.POST.get('branch')
                justification = request.POST.get('justification', '').strip()
                priority = request.POST.get('priority', ApprovalRequest.PRIORITY_MEDIUM)
                
                # Validate required fields
                # NOTE: The unified /assets/register/ form does not expose a
                # dedicated "title" field for managers; we derive a sensible
                # title later from the category/description. Only category and
                # branch are strictly required here.
                if not all([category_id, branch_id]):
                    messages.error(request, "Please fill in all required fields.")
                    context = self.get_context_data()
                    return self.render_to_response(context)
                
                # Validate category
                try:
                    category = AssetCategory.objects.for_company(company).get(pk=category_id)
                except AssetCategory.DoesNotExist:
                    messages.error(request, "Invalid category selected.")
                    context = self.get_context_data()
                    return self.render_to_response(context)
                
                # Validate branch
                try:
                    branch = Branch.objects.get(pk=branch_id, company=company, is_active=True)
                except Branch.DoesNotExist:
                    messages.error(request, "Invalid branch selected.")
                    context = self.get_context_data()
                    return self.render_to_response(context)
                
                # Build asset data
                asset_data = {
                    'category_id': int(category_id),
                    'branch_id': int(branch_id),
                    'description': description,
                    'status': request.POST.get('status', 'active'),
                    'dynamic_data': {},
                }
                
                # Extract dynamic fields
                fields = AssetCategoryField.objects.for_company(company).filter(category=category)
                for field in fields:
                    # Support both legacy "field_" prefix (used by the
                    # dedicated asset-creation request form) and the
                    # "dyn_" prefix used by the unified asset registration
                    # form and AssetForm/asset_form.js.
                    raw_value = request.POST.get(
                        f'dyn_{field.key}',
                        request.POST.get(f'field_{field.key}', '')
                    )
                    field_value = (raw_value or '').strip()
                    if field_value:
                        asset_data['dynamic_data'][field.key] = field_value
                    elif field.required:
                        messages.error(request, f"Required field '{field.label}' is missing.")
                        context = self.get_context_data()
                        return self.render_to_response(context)
                
                # Check for assigned_to
                assigned_to_id = request.POST.get('assigned_to_id')
                if assigned_to_id:
                    asset_data['assigned_to_id'] = int(assigned_to_id)
                
                # Create approval request
                approval_request = ApprovalRequest.objects.create(
                    company=company,
                    branch=branch,
                    request_type=ApprovalRequest.TYPE_ASSET_CREATION,
                    # Derive a clear, human-readable title when the form
                    # does not supply one explicitly.
                    title=title or f"Asset Creation: {category.name}",
                    description=f"{description}\n\nJustification: {justification}",
                    requested_by=user,
                    priority=priority,
                    metadata={
                        'asset_data': asset_data,
                        'justification': justification,
                        'category_name': category.name,
                    }
                )
                
                # Auto-assign to admin
                from django.contrib.auth import get_user_model
                User = get_user_model()
                admin = User.objects.filter(
                    company=company,
                    role='admin',
                    is_active=True
                ).first()
                
                if admin:
                    approval_request.assigned_to = admin
                    approval_request.save()
                    
                    # Create notification
                    Alert.objects.create(
                        company=company,
                        branch=branch,
                        recipient=admin,
                        level=Alert.LEVEL_INFO,
                        message=f"New asset creation request from {user.get_full_name() or user.username}: {title}",
                        context={
                            'request_id': approval_request.pk,
                            'request_type': 'asset_creation',
                            'requested_by': user.pk,
                        }
                    )
                
                # Log audit
                log_audit(
                    user,
                    "asset_creation_requested",
                    details=f"Requested asset creation: {title}",
                    company=company,
                    branch=branch,
                    metadata={
                        'request_id': approval_request.pk,
                        'category': category.name,
                        'priority': priority,
                    }
                )
                
                messages.success(
                    request,
                    f"Asset creation request '{title}' submitted successfully. "
                    f"Assigned to {approval_request.assigned_to.get_full_name() if approval_request.assigned_to else 'admin'} for review."
                )
                
                return redirect('approval_dashboard')
        
        except Exception as e:
            messages.error(request, f"Failed to submit request: {str(e)}")
            context = self.get_context_data()
            return self.render_to_response(context)


asset_create = user_passes_test(is_admin_or_manager, login_url='users:login')(AssetCreateView.as_view())

# Time Stamp For File Names
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

# Asset update view for admin/manager with audit logging
class AssetUpdateView(UserPassesTestMixin, UpdateView):
    model = Asset
    form_class = AssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('asset_list')
    slug_field = 'uuid'
    slug_url_kwarg = 'uuid'

    def test_func(self):
        user = self.request.user
        return user.is_authenticated and can(user, 'edit_assets')
    
    def get_queryset(self):
        """
        WORLD-CLASS: Apply company-scoped filtering for asset updates.
        
        Only admins and managers can edit assets (enforced by test_func).
        Both roles can edit any company asset without branch restrictions.
        
        Security:
        - Must have company context
        - Admins: Can edit all company assets
        - Managers: Can edit all company assets
        - Users: Blocked by test_func (can't edit assets)
        """
        user = self.request.user
        company = getattr(self.request, 'company', None)
        role = getattr(user, 'role', 'user')
        
        if not company:
            # No company context: no access
            return Asset.objects.none()
        
        # Filter by company
        qs = Asset.objects.filter(company=company)
        
        # WORLD-CLASS: Admins and managers can edit any company asset
        # No branch restriction for edit functionality
        if role in ('admin', 'manager'):
            pass  # No additional filtering
        else:
            # Safety fallback: users shouldn't reach here due to test_func
            # But if they do, give them no access
            qs = qs.none()
        
        return qs
    
    def get_object(self, queryset=None):
        """
        WORLD-CLASS FIX: Ensure dynamic_data integrity before editing.
        
        Problem Solved:
        - Assets created before category fields were added have incomplete dynamic_data
        - Missing keys cause pre-fill to fail on edit page
        - Users see empty fields that should show as editable
        
        Solution:
        - Call ensure_dynamic_data_integrity() before returning object
        - This initializes missing keys with appropriate defaults
        - Pre-fill will now work correctly for all fields
        
        Performance: O(1) - single query + O(n) field initialization (n = 5-20 typically)
        """
        obj = super().get_object(queryset)
        
        # CRITICAL: Ensure all category fields exist in dynamic_data
        # This fixes pre-fill failures for assets with incomplete data
        if obj.ensure_dynamic_data_integrity():
            # Save if data was updated (prevents data loss)
            obj.save(update_fields=['dynamic_data'])
            print(f"✅ INTEGRITY FIX: Updated dynamic_data for asset {obj.uuid}")
        
        return obj

    def form_valid(self, form):
        """
        WORLD-CLASS IMPLEMENTATION: Handle asset updates with proper data persistence.
        
        Architecture:
        1. Capture old state for audit trail
        2. Detect status changes vs normal edits
        3. Route to appropriate handler (status service or direct save)
        4. Single save point - no double saves
        5. Complete audit logging
        6. Clear success messages
        
        Multi-tenancy: All operations scoped to user's company
        Security: Permissions checked, all changes audited
        Performance: Single transaction, optimized queries
        """
        from decimal import Decimal
        from assets.services.status_changes import AssetStatusChangeService
        from django.contrib import messages
        from django.db import transaction
        
        # DEBUG: Log form submission
        print(f"✅ FORM_VALID called for asset {self.get_object().uuid}")
        print(f"✅ Form data: {form.cleaned_data.keys()}")
        
        # Capture original state from database for audit trail
        old_obj = self.get_object()
        old_assigned = old_obj.assigned_to
        old_status = old_obj.status
        old_branch = old_obj.branch
        try:
            old_dyn = dict(old_obj.dynamic_data or {})
        except Exception:
            old_dyn = {}
        
        # Get modified asset from form (has all form changes in memory)
        asset = form.instance
        
        # Detect status change
        status_changed = form.cleaned_data.get('_status_changed', False)
        new_status = form.cleaned_data.get('_new_status')
        
        # Use atomic transaction for data consistency
        try:
            with transaction.atomic():
                if status_changed and new_status:
                    # ============================================================
                    # STATUS CHANGE WORKFLOW
                    # ============================================================
                    reason = form.cleaned_data.get('status_change_reason', '')
                    
                    # CRITICAL FIX: Prepare asset with all form changes BUT preserve original status
                    # The service layer will handle the status change itself
                    asset = form.save(commit=False)
                    
                    # Restore original status - service will change it properly
                    # This is critical because services validate the current status
                    asset.status = old_status
                    
                    # Route to appropriate status change service
                    if new_status == 'in_maintenance':
                        maintenance_type = form.cleaned_data.get('maintenance_type', 'corrective')
                        asset, maintenance_record = AssetStatusChangeService.change_to_in_maintenance(
                            asset=asset,
                            user=self.request.user,
                            reason=reason,
                            maintenance_type=maintenance_type
                        )
                        messages.success(
                            self.request,
                            f'✅ Asset placed under {maintenance_type} maintenance. '
                            f'Maintenance record #{maintenance_record.id} created.'
                        )
                    
                    elif new_status == 'retired':
                        disposal_method = form.cleaned_data.get('disposal_method')
                        asset = AssetStatusChangeService.change_to_retired(
                            asset=asset,
                            user=self.request.user,
                            reason=reason,
                            disposal_method=disposal_method,
                            salvage_value=None
                        )
                        messages.warning(
                            self.request,
                            f'⚠️ Asset retired. Disposal method: {disposal_method}.'
                        )
                    
                    elif new_status == 'lost':
                        loss_date = form.cleaned_data.get('loss_date')
                        loss_reason = form.cleaned_data.get('loss_reason')
                        loss_details = form.cleaned_data.get('loss_details', '')
                        last_known_location = form.cleaned_data.get('last_known_location', '')
                        police_report = form.cleaned_data.get('police_report_number', '')
                        
                        asset = AssetStatusChangeService.change_to_lost(
                            asset=asset,
                            user=self.request.user,
                            loss_date=loss_date,
                            loss_reason=loss_reason,
                            details=loss_details,
                            last_known_location=last_known_location,
                            police_report_number=police_report
                        )
                        messages.error(
                            self.request,
                            f'🚨 Asset reported {loss_reason}. High-priority alert created.'
                        )
                    
                    else:
                        # Generic status change (active, deleted, etc.)
                        asset.status = new_status
                        asset.status_changed_at = timezone.now()
                        asset.status_changed_by = self.request.user
                        asset.status_change_reason = reason
                        asset.save()
                        messages.success(
                            self.request,
                            f'✅ Asset status changed to {new_status.replace("_", " ").title()}.'
                        )
                    
                    # Save many-to-many relationships
                    form.save_m2m()
                    
                    # Log status change
                    log_audit(
                        self.request.user,
                        'status_change',
                        asset,
                        f'Status changed: {old_status} → {new_status}. Reason: {reason[:50]}'
                    )
                    
                else:
                    # ============================================================
                    # NORMAL EDIT WORKFLOW (No status change)
                    # ============================================================
                    
                    # DEBUG: Log normal edit path
                    print(f"📝 NORMAL EDIT: Saving asset without status change")
                    print(f"📝 assigned_to in form: {form.cleaned_data.get('assigned_to')}")
                    print(f"📝 branch in form: {form.cleaned_data.get('branch')}")
                    
                    # Save asset with all form changes
                    asset = form.save()
                    
                    # DEBUG: Verify save
                    print(f"✅ SAVED: assigned_to = {asset.assigned_to}")
                    print(f"✅ SAVED: branch = {asset.branch}")
                    print(f"✅ SAVED: dynamic_data = {asset.dynamic_data}")
                    
                    # Build change summary for user feedback
                    changes = []
                    
                    # Track assignment changes
                    new_assigned = asset.assigned_to
                    if old_assigned != new_assigned:
                        if new_assigned:
                            log_audit(
                                self.request.user,
                                ASSIGN_ACTION,
                                asset,
                                f'Asset assigned to {new_assigned}',
                                related_user=new_assigned
                            )
                            changes.append(f'assigned to {new_assigned.get_full_name() or new_assigned.username}')
                        else:
                            log_audit(
                                self.request.user,
                                'unassign',
                                asset,
                                f'Asset unassigned (previously {old_assigned})'
                            )
                            changes.append('unassigned')
                    
                    # Track branch changes
                    new_branch = asset.branch
                    if old_branch != new_branch:
                        log_audit(
                            self.request.user,
                            'branch_change',
                            asset,
                            f'Branch changed: {old_branch} → {new_branch}'
                        )
                        changes.append(f'moved to {new_branch.name if new_branch else "head office"}')
                    
                    # Track dynamic data changes
                    try:
                        new_dyn = dict(asset.dynamic_data or {})
                        changed_keys = [
                            k for k in set(old_dyn.keys()) | set(new_dyn.keys())
                            if old_dyn.get(k) != new_dyn.get(k)
                        ]
                        
                        if changed_keys:
                            # Log warranty changes specifically
                            if 'warranty_expiry' in changed_keys or 'warranty_provider' in changed_keys:
                                log_audit(
                                    self.request.user,
                                    'warranty_change',
                                    asset,
                                    f"Warranty updated: Provider='{new_dyn.get('warranty_provider', 'N/A')}', "
                                    f"Expiry='{new_dyn.get('warranty_expiry', 'N/A')}'"
                                )
                                changes.append('warranty information updated')
                            
                            # Log other dynamic field changes
                            preview = ', '.join(changed_keys[:5])
                            if len(changed_keys) > 5:
                                preview += f' and {len(changed_keys) - 5} more'
                            log_audit(
                                self.request.user,
                                'edit',
                                asset,
                                f'Asset details updated: {preview}'
                            )
                    except Exception as e:
                        # Don't fail on audit logging errors
                        pass
                    
                    # Show success message with change summary
                    if changes:
                        change_summary = ', '.join(changes)
                        messages.success(
                            self.request,
                            f'✅ Asset updated successfully: {change_summary}.'
                        )
                    else:
                        messages.success(self.request, '✅ Asset updated successfully.')
                
                # Transaction committed here
                
        except (PermissionDenied, ValidationError) as e:
            messages.error(self.request, f'❌ Error: {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ Unexpected error: {str(e)}')
            return self.form_invalid(form)
        
        # CRITICAL: Refresh asset from database to ensure all changes are committed
        # This prevents stale data issues when redirecting
        asset.refresh_from_db()
        
        # DEBUG: Verify final state
        print(f"✅ FINAL STATE: status={asset.status}, assigned_to={asset.assigned_to}")
        
        # Return redirect to success URL with cache-busting
        response = HttpResponseRedirect(self.get_success_url())
        # Prevent browser caching of the redirect to ensure fresh data
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    
    def form_invalid(self, form):
        """
        WORLD-CLASS ERROR HANDLING: Provide detailed feedback on validation errors.
        """
        from django.contrib import messages
        
        # Log form errors for debugging
        error_details = []
        
        # Field-specific errors
        for field, errors in form.errors.items():
            if field == '__all__':
                for error in errors:
                    error_details.append(f"Form error: {error}")
            else:
                for error in errors:
                    error_details.append(f"{field}: {error}")
        
        # Show detailed error message to user
        if error_details:
            error_msg = "❌ Validation errors: " + "; ".join(error_details[:3])
            if len(error_details) > 3:
                error_msg += f" and {len(error_details) - 3} more"
            messages.error(self.request, error_msg)
        else:
            messages.error(self.request, "❌ Please correct the errors in the form.")
        
        # Log to console for debugging
        print(f"🔴 FORM VALIDATION FAILED for asset {self.get_object().uuid}")
        print(f"🔴 Errors: {form.errors}")
        if hasattr(form, 'cleaned_data'):
            print(f"🔴 Cleaned data: {form.cleaned_data}")
        
        return super().form_invalid(form)

    def get_initial(self):
        """
        WORLD-CLASS: Pre-populate ALL form fields with current asset data.
        
        This ensures a seamless edit experience where users see all existing data
        without having to re-enter anything. Matches ServiceNow ITAM, IBM Maximo,
        and SAP EAM standards for edit forms.
        
        Pre-populates:
        - Standard fields (company, category, branch, status, assigned_to, description)
        - Maintenance fields (maintenance_enabled, maintenance_interval_days, maintenance_notes)
        - Dynamic category-specific fields from asset.dynamic_data
        - File fields (images, documents, qr_code)
        """
        initial = super().get_initial()
        asset = self.get_object()
        
        # ============================================================
        # CRITICAL FIELDS (Multi-Tenancy & Security)
        # ============================================================
        
        # Pre-populate company (required for validation)
        if asset.company:
            initial['company'] = asset.company.id
        
        # Pre-populate category (triggers dynamic field loading)
        if asset.category:
            initial['category'] = asset.category.id
        
        # Pre-populate branch field (critical for multi-tenancy UX)
        if asset.branch:
            initial['branch'] = asset.branch.id
        
        # Pre-populate status (current state)
        if asset.status:
            initial['status'] = asset.status
        
        # Pre-populate assigned_to field (UX improvement)
        if asset.assigned_to:
            initial['assigned_to'] = asset.assigned_to.id
        
        # ============================================================
        # STANDARD FIELDS
        # ============================================================
        
        # Pre-populate description
        if asset.description:
            initial['description'] = asset.description
        
        # ============================================================
        # ============================================================
        # MAINTENANCE FIELDS (Preventive Maintenance)
        # ============================================================
        
        # Pre-populate maintenance tracking fields
        initial['maintenance_enabled'] = asset.maintenance_enabled
        
        if asset.maintenance_interval_days:
            initial['maintenance_interval_days'] = asset.maintenance_interval_days
        
        if asset.maintenance_notes:
            initial['maintenance_notes'] = asset.maintenance_notes
        
        # ============================================================
        # DYNAMIC CATEGORY-SPECIFIC FIELDS
        # ============================================================
        
        # Pre-populate dynamic fields from asset.dynamic_data JSON
        # This is CRITICAL for world-class UX - users should see all their data
        if asset.dynamic_data:
            try:
                dynamic_data = asset.dynamic_data if isinstance(asset.dynamic_data, dict) else {}
                
                # Populate each dynamic field with dyn_ prefix
                for key, value in dynamic_data.items():
                    field_name = f'dyn_{key}'
                    
                    # Only populate if value exists and is not None
                    if value is not None and value != '':
                        initial[field_name] = value
                        
                print(f"✅ PRE-POPULATED {len(dynamic_data)} dynamic fields for asset {asset.uuid}")
                
            except Exception as e:
                # Don't fail the entire form if dynamic data has issues
                print(f"⚠️ Warning: Could not pre-populate dynamic fields: {e}")
        
        # ============================================================
        # FILE FIELDS (Images, Documents, QR Codes)
        # ============================================================
        # Note: File fields are handled differently - they show current file with option to replace
        # The template handles display of existing files
        
        print(f"✅ EDIT FORM: Pre-populated {len(initial)} fields for asset {asset.uuid}")
        
        return initial

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Mirror CreateView: pass request for permission-aware form behavior
        kwargs['initial'] = kwargs.get('initial', {})
        kwargs['initial']['request'] = self.request
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Provide user role to template/JS (e.g., to show admin-only controls)
        context['user_role'] = getattr(self.request.user, 'role', 'user')
        # Add current asset for context display
        context['current_asset'] = self.get_object()
        return context

@require_GET
def get_dynamic_fields(request):
    category_id = request.GET.get('category_id')
    try:
        # Source fields from AssetCategoryField records to avoid relying on cached JSON
        company = getattr(request, 'company', None)
        if not company:
            return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
        fields_qs = AssetCategoryField.objects.for_company(company).filter(category_id=category_id)
        schema = {}
        for f in fields_qs:
            schema[f.key] = {
                'type': f.type,
                'label': f.label,
                'required': f.required,
            }
        return JsonResponse({'success': True, 'fields': schema})
    except Exception:
        return JsonResponse({'success': False, 'fields': {}})

# Asset list view: only for authenticated users
class AssetListView(CompanyScopedQuerysetMixin, BranchContextMixin, LoginRequiredMixin, ListView):
    model = Asset
    template_name = 'assets/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 20

    def get_paginate_by(self, queryset):
        """Allow dynamic page size from query params (world-class UX)."""
        page_size = self.request.GET.get('page_size', 20)
        try:
            page_size = int(page_size)
            # Limit to reasonable values for performance
            if page_size > 200:
                page_size = 200
            elif page_size < 10:
                page_size = 10
            return page_size
        except (ValueError, TypeError):
            return 20

    def get_queryset(self):
        """
        WORLD-CLASS: Use unified filtering service for consistency.
        
        This ensures the asset list matches statistics exactly.
        All role-based filtering is handled by the service.
        """
        from assets.services.filtering import asset_filtering_service
        
        company = getattr(self.request, 'company', None)
        user = self.request.user
        
        # Get base queryset using unified filtering service
        # This handles role-based filtering (admin/manager/user) consistently
        qs = asset_filtering_service.get_base_queryset(company, user, self.request)
        
        # Extract filter parameters from request
        category = self.request.GET.get('category')
        status = self.request.GET.get('status')
        location = self.request.GET.get('location')
        search = self.request.GET.get('search')
        assigned = self.request.GET.get('assigned')
        warranty = self.request.GET.get('warranty')
        branch = self.request.GET.get('branch')  # Multi-tenancy: branch filter
        
        # Apply standard filters using the service
        qs = asset_filtering_service.apply_status_filter(qs, status)
        qs = asset_filtering_service.apply_branch_filter(qs, branch)
        qs = asset_filtering_service.apply_category_filter(qs, category)
        qs = asset_filtering_service.apply_assigned_filter(qs, assigned)
        qs = asset_filtering_service.apply_warranty_filter(qs, warranty)
        qs = asset_filtering_service.apply_location_filter(qs, location)
        qs = asset_filtering_service.apply_search_filter(qs, search)
        
        # Dynamic field filters (category-specific)
        # These are applied AFTER standard filters for category-specific fields
        if category:
            # Fetch dynamic fields for this category
            fields = AssetCategoryField.objects.for_company(company).filter(category_id=category)
            for field in fields:
                val = self.request.GET.get(f'dyn_{field.key}')
                if val:
                    # Filter by dynamic_data JSON key
                    if field.type == 'text':
                        qs = qs.filter(**{f'dynamic_data__{field.key}__icontains': val})
                    elif field.type == 'number':
                        try:
                            qs = qs.filter(**{f'dynamic_data__{field.key}': float(val)})
                        except ValueError:
                            pass
                    elif field.type == 'date':
                        # Parse mm/dd/yyyy and convert to yyyy-MM-dd
                        import datetime
                        try:
                            dt = datetime.datetime.strptime(val, '%m/%d/%Y')
                            iso_val = dt.strftime('%Y-%m-%d')
                            qs = qs.filter(**{f'dynamic_data__{field.key}': iso_val})
                        except ValueError:
                            pass  # Invalid date format, ignore filter
        
        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = getattr(self.request, 'company', None)
        user = self.request.user
        role = getattr(user, 'role', 'user') if hasattr(user, 'role') else 'user'
        
        # Get company-scoped categories with asset counts (only show categories with assets)
        if company:
            context['categories'] = AssetCategory.objects.for_company(company).annotate(
                asset_count=Count('assets', filter=Q(assets__company=company))
            ).filter(asset_count__gt=0).order_by('name')
        else:
            context['categories'] = AssetCategory.objects.none()
        
        context['statuses'] = Asset.STATUS_CHOICES
        
        # Get branches for filtering (multi-tenancy)
        try:
            from tenancy.models import Branch
            if company:
                context['branches'] = Branch.objects.filter(company=company, is_active=True).order_by('name')
            else:
                context['branches'] = []
        except (ImportError, AttributeError):
            context['branches'] = []
        
        # WORLD-CLASS FIX: Use unified filtering service for statistics
        # This ensures statistics match dashboard metrics EXACTLY
        try:
            from assets.services.filtering import asset_filtering_service
            stats = asset_filtering_service.get_statistics(company, user, self.request)
            context['statistics'] = {
                'total': stats['total'],
                'active': stats['active'],
                'in_maintenance': stats['in_maintenance'],
                'retired': stats['retired'],
                'unassigned': stats['unassigned'],
                'assigned': stats['assigned'],
            }
        except Exception as e:
            # Fallback statistics if there's an error
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error calculating statistics: {e}")
            context['statistics'] = {
                'total': 0,
                'active': 0,
                'in_maintenance': 0,
                'retired': 0,
                'unassigned': 0,
                'assigned': 0,
            }
        
        # Get current branch from session or user
        current_branch = getattr(self.request, 'branch', None) or getattr(user, 'primary_branch', None)
        context['current_branch'] = current_branch
        context['user_role'] = role
        context['is_admin'] = role == 'admin'
        context['is_manager'] = role == 'manager'
        
        # Dynamic fields logic
        selected_category = self.request.GET.get('category')
        if selected_category:
            context['dynamic_fields'] = AssetCategoryField.objects.for_company(company).filter(category_id=selected_category)
        else:
            all_fields = AssetCategoryField.objects.for_company(company)
            seen = set()
            unique_fields = []
            for f in all_fields:
                dedup_key = (f.key.lower().strip(), f.label.lower().strip())
                if dedup_key not in seen:
                    unique_fields.append(f)
                    seen.add(dedup_key)
            context['dynamic_fields'] = unique_fields
        
        return context
    
    def render_to_response(self, context, **response_kwargs):
        """
        WORLD-CLASS: Add cache-busting headers to ensure fresh data.
        Prevents browser from caching stale asset data after edits.
        """
        response = super().render_to_response(context, **response_kwargs)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

class AssetDetailView(BranchContextMixin, CompanyScopedQuerysetMixin, LoginRequiredMixin, DetailView):
    model = Asset
    template_name = 'assets/asset_detail.html'
    context_object_name = 'asset'
    
    def get_queryset(self):
        """
        WORLD-CLASS FIX: Asset detail view should not apply branch filtering for admins.
        
        Detail views are for viewing SPECIFIC assets by ID, so:
        - Admins: See all company assets (no branch restriction)
        - Managers: See all company assets (branch restriction would be too limiting)
        - Users: Handled by CompanyScopedQuerysetMixin (shows only assigned assets)
        
        This ensures admins can view any asset detail page, regardless of
        currently selected branch in the session.
        """
        company = self.get_company()
        user = self.request.user
        role = getattr(user, 'role', 'user')
        
        # Get base queryset scoped to company
        qs = Asset.objects.filter(company=company)
        
        # WORLD-CLASS: Role-based filtering for detail view
        if role == 'admin':
            # Admins see ALL company assets (no branch filtering)
            pass
        elif role == 'manager':
            # Managers see ALL company assets (no branch filtering for detail views)
            # This allows managers to view assets they may need to manage
            pass
        elif role == 'user':
            # Users see only assets assigned to them
            # SECURITY: This prevents users from viewing other users' assets
            qs = qs.filter(assigned_to=user)
        else:
            # Fallback for unknown roles: no access
            qs = qs.none()
        
        return qs

    def get(self, request, *args, **kwargs):
        asset = self.get_object()
        log_audit(request.user, 'view', asset, 'Asset viewed via dashboard')
        # Redirect to UUID-based URL if accessed by PK
        return redirect('asset_detail_by_uuid', uuid=asset.uuid)

class AssetScanView(TemplateView):
    template_name = 'assets/asset_scan_enterprise.html'

@require_GET
@csrf_exempt
def asset_by_code(request):
    code = request.GET.get('code')
    asset = None
    if code:
        # Try to extract UUID from QR code format 'ASSET|v1|{uuid}'
        match = re.match(r'^ASSET\|v1\|([0-9a-fA-F-]{36})$', code)
        if match:
            uuid_val = match.group(1)
            asset = Asset.objects.filter(uuid=uuid_val).first()
        # Fallback: Try by QR code filename or asset ID
        if not asset:
            asset = Asset.objects.filter(qr_code__icontains=code).first()
        if not asset and code.isdigit():
            asset = Asset.objects.filter(pk=int(code)).first()
    if asset:
        # Log QR code scan (unauthenticated permitted here)
        log_audit(request.user if request.user.is_authenticated else None, 'scan', asset, f'QR code scanned: {code}')
        # Return only non-sensitive, minimal fields suitable for public display
        safe_dynamic = {
            'name': asset.dynamic_data.get('name'),
            'model': asset.dynamic_data.get('model'),
        }
        data = {
            'id': asset.pk,
            'dynamic_data': safe_dynamic,
            'category_name': asset.category.name,
            'status': asset.status,
        }
        return JsonResponse({'success': True, 'asset': data})
    return JsonResponse({'success': False})

@method_decorator(ensure_csrf_cookie, name='dispatch')
class AssetDetailByUUIDView(DetailView):
    """
    WORLD-CLASS: Asset detail view with dual-mode access (public/authenticated)
    
    @ensure_csrf_cookie decorator ensures CSRF cookie is set for AJAX requests
    This is critical for transfer functionality to work correctly
    
    Security Model:
    - Public (unauthenticated): Can view basic asset info for QR code scanning
    - Authenticated users: Must have company access
      * Admins: Can view all company assets (no branch restriction)
      * Managers: Can view all company assets (no branch restriction)
      * Users: Can view only assets assigned to them
    """
    model = Asset
    template_name = 'assets/asset_detail_enhanced.html'
    context_object_name = 'asset'
    slug_field = 'uuid'
    slug_url_kwarg = 'uuid'
    
    def get_queryset(self):
        """
        WORLD-CLASS: Apply company-scoped and role-based filtering.
        
        Public users can view any asset (for QR code scanning use case).
        Authenticated users must have company access and appropriate permissions.
        """
        qs = Asset.objects.all()
        
        # If user is authenticated, apply security filters
        user = self.request.user
        if user.is_authenticated:
            company = getattr(self.request, 'company', None)
            role = getattr(user, 'role', 'user')
            
            # SECURITY: Must have company context
            if company:
                # Filter by company first
                qs = qs.filter(company=company)
                
                # WORLD-CLASS: Role-based filtering
                if role == 'admin':
                    # Admins see ALL company assets (no branch restriction)
                    pass
                elif role == 'manager':
                    # Managers see ALL company assets (no branch restriction)
                    # This allows cross-branch visibility for management
                    pass
                elif role == 'user':
                    # Users see only assets assigned to them
                    # SECURITY: Prevents users from viewing other users' assets
                    qs = qs.filter(assigned_to=user)
                else:
                    # Unknown role: no access
                    qs = qs.none()
            else:
                # Authenticated but no company context: no access
                # This handles edge cases with misconfigured accounts
                qs = qs.none()
        # else: Unauthenticated users can view any asset (public QR scanning)
        
        return qs

    def get_device_info(self, user_agent):
        """Extract device information from user agent."""
        if 'Android' in user_agent or 'iPhone' in user_agent or 'iPad' in user_agent:
            return 'Mobile App/Browser'
        elif 'Windows' in user_agent or 'Macintosh' in user_agent or 'Linux' in user_agent:
            return 'Desktop/Scanner Device'
        elif 'ZBar' in user_agent or 'Scanner' in user_agent:
            return 'Hardware QR Scanner'
        return 'Unknown'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        asset = self.object
        user = self.request.user
        is_authenticated = user.is_authenticated
        
        # Determine access level (safely handle anonymous users)
        role = user.role.lower() if is_authenticated and hasattr(user, 'role') and user.role else ''
        is_admin = role == 'admin'
        is_manager = role == 'manager'
        is_user = role == 'user'
        
        # Check if user has access to this asset's company
        has_company_access = False
        if is_authenticated and hasattr(user, 'company_id'):
            has_company_access = user.company_id == asset.company_id
        
        # Public view context (minimal, safe data)
        context['is_public_view'] = not is_authenticated
        context['is_authenticated'] = is_authenticated
        context['has_company_access'] = has_company_access
        context['is_admin'] = is_admin
        context['is_manager'] = is_manager
        context['is_user'] = is_user
        
        # WORLD-CLASS: Smart transfer permissions based on ownership and role
        # Following ServiceNow ITAM, IBM Maximo, and SAP EAM best practices:
        # - Admins: Can transfer any asset in their company
        # - Managers: Can transfer assets in their managed branches
        # - Users: Can transfer assets assigned to them OR unassigned assets in their branches
        can_transfer = False
        if is_authenticated and has_company_access:
            if is_admin:
                # Admins can transfer any company asset
                can_transfer = True
            elif is_manager:
                # Managers can transfer assets in their managed branches
                can_transfer = True
            elif is_user:
                # Users can transfer:
                # 1. Assets assigned to them (ownership-based)
                # 2. Unassigned assets in their accessible branches (branch-based)
                is_assigned_to_user = asset.assigned_to_id == user.pk
                is_unassigned_in_accessible_branch = False
                
                if not asset.assigned_to_id:
                    # Check if asset is in user's accessible branches
                    try:
                        from tenancy.policy_service import PolicyService
                        accessible_branch_ids = PolicyService.get_accessible_branches(user, asset.company)
                        if accessible_branch_ids:
                            is_unassigned_in_accessible_branch = asset.branch_id in accessible_branch_ids
                        else:
                            # No accessible branches, check primary branch
                            is_unassigned_in_accessible_branch = (
                                hasattr(user, 'primary_branch') and 
                                user.primary_branch and 
                                asset.branch_id == user.primary_branch.id
                            )
                    except Exception as e:
                        # Fallback: check if asset is in user's primary branch
                        is_unassigned_in_accessible_branch = (
                            hasattr(user, 'primary_branch') and 
                            user.primary_branch and 
                            asset.branch_id == user.primary_branch.id
                        )
                
                can_transfer = is_assigned_to_user or is_unassigned_in_accessible_branch
        
        # Role-based permissions
        context['can_view_sensitive'] = is_authenticated and has_company_access
        context['can_edit'] = (is_admin or is_manager) and has_company_access
        context['can_transfer'] = can_transfer
        context['can_delete'] = is_admin and has_company_access
        
        # Public data (always visible) with safe access
        try:
            dynamic_data = asset.dynamic_data or {}
            name = dynamic_data.get('name', asset.category.name) if dynamic_data else asset.category.name
        except Exception:
            name = asset.category.name
        
        context['public_data'] = {
            'name': name,
            'category': asset.category.name if asset.category else 'Unknown',
            'status': asset.get_status_display(),
            'status_badge_class': self.get_status_badge_class(asset.status),
        }
        
        if is_authenticated and has_company_access:
            # WORLD-CLASS: Policy-driven user filtering for transfers
            from users.models import User
            from tenancy.policy_service import PolicyService
            
            # Get accessible users based on role and policy
            users_qs = User.objects.filter(
                company_id=asset.company_id,
                is_active=True
            ).exclude(pk=user.pk).select_related('company')
            
            # Apply policy-driven branch scoping with robust error handling
            try:
                if role == 'user':
                    # Users can only transfer to users in their accessible branches
                    try:
                        accessible_branch_ids = PolicyService.get_accessible_branches(user, asset.company)
                        if accessible_branch_ids:
                            users_qs = users_qs.filter(
                                user_branches__branch_id__in=accessible_branch_ids
                            ).distinct()
                        else:
                            # No accessible branches, show only self
                            users_qs = User.objects.none()
                    except Exception as e:
                        # Fallback: only users in same branch
                        if hasattr(user, 'primary_branch') and user.primary_branch:
                            users_qs = users_qs.filter(
                                user_branches__branch=user.primary_branch
                            ).distinct()
                        else:
                            # No primary branch, show only self
                            users_qs = User.objects.none()
                elif role == 'manager':
                    # Managers can transfer to users in their accessible branches
                    try:
                        accessible_branch_ids = PolicyService.get_accessible_branches(user, asset.company)
                        if accessible_branch_ids:
                            users_qs = users_qs.filter(
                                user_branches__branch_id__in=accessible_branch_ids
                            ).distinct()
                        # If no accessible branches, return all company users
                    except Exception as e:
                        # Fallback: all company users
                        pass
                # Admin sees all company users (no additional filter)
            except Exception as e:
                # If everything fails, provide basic queryset
                users_qs = User.objects.filter(
                    company_id=asset.company_id,
                    is_active=True
                ).exclude(pk=user.pk)
            
            context['available_users'] = users_qs.order_by('first_name', 'last_name')[:100]
            
            # WORLD-CLASS: Policy-driven branch filtering for transfers
            from tenancy.models import Branch
            from tenancy.policy_service import PolicyService
            
            # Get accessible branches based on role and policy
            branches_qs = Branch.objects.filter(
                company_id=asset.company_id,
                is_active=True
            )
            
            # Apply policy-driven branch scoping with error handling
            if role in ('user', 'manager'):
                # Users and managers see only accessible branches
                try:
                    accessible_branch_ids = PolicyService.get_accessible_branches(user, asset.company)
                    if accessible_branch_ids:
                        branches_qs = branches_qs.filter(id__in=accessible_branch_ids)
                    else:
                        # No accessible branches, show current branch if available
                        if hasattr(user, 'primary_branch') and user.primary_branch:
                            branches_qs = branches_qs.filter(id=user.primary_branch.id)
                        else:
                            # No branches available
                            branches_qs = Branch.objects.none()
                except Exception as e:
                    # Fallback: current branch only
                    if hasattr(user, 'primary_branch') and user.primary_branch:
                        branches_qs = branches_qs.filter(id=user.primary_branch.id)
                    else:
                        # No branch fallback
                        branches_qs = Branch.objects.none()
            # Admin sees all company branches (no additional filter)
            
            context['available_branches'] = branches_qs.order_by('name')
            
            # WORLD-CLASS: Check for pending transfer on this asset
            from assets.models import AssetTransfer
            pending_transfer = AssetTransfer.objects.filter(
                asset=asset,
                state__in=AssetTransfer.ACTIVE_STATES
            ).select_related('to_user', 'from_user', 'initiator').first()
            context['asset'].pending_transfer = pending_transfer
            
            # Transfer history (last 10)
            context['transfer_history'] = asset.transfers.select_related(
                'from_user', 'to_user', 'from_branch', 'to_branch', 'approved_by'
            ).order_by('-created_at')[:10]
            
            # Maintenance records (last 10)
            context['maintenance_records'] = asset.maintenance_records.select_related(
                'performed_by', 'supervisor', 'created_by'
            ).order_by('-scheduled_for')[:10]
            
            # Audit events (last 20)
            context['audit_events'] = asset.auditlog_set.select_related(
                'user', 'branch'
            ).order_by('-timestamp')[:20]
            
            # Related assets (same category, same branch)
            context['related_assets'] = Asset.objects.filter(
                category=asset.category,
                branch=asset.branch,
                company=asset.company
            ).exclude(pk=asset.pk).select_related('category', 'branch')[:5]
            
            # Maintenance status with error handling
            try:
                if asset.maintenance_enabled and asset.next_maintenance_date:
                    from datetime import date, timedelta
                    days_until_maintenance = (asset.next_maintenance_date - date.today()).days
                    context['maintenance_status'] = {
                        'enabled': True,
                        'next_date': asset.next_maintenance_date,
                        'days_until': days_until_maintenance,
                        'days_overdue': abs(days_until_maintenance) if days_until_maintenance < 0 else 0,
                        'is_overdue': days_until_maintenance < 0,
                        'is_due_soon': 0 <= days_until_maintenance <= 7,
                        'last_date': asset.last_maintenance_date,
                    }
                else:
                    context['maintenance_status'] = {
                        'enabled': False,
                    }
            except Exception as e:
                context['maintenance_status'] = {
                    'enabled': False,
                    'error': True,
                }
            
            # Asset utilization (based on audit logs) with error handling
            try:
                from django.db.models import Count
                from datetime import datetime, timedelta
                thirty_days_ago = datetime.now() - timedelta(days=30)
                context['utilization_stats'] = {
                    'scans_30_days': asset.auditlog_set.filter(
                        action='scan',
                        timestamp__gte=thirty_days_ago
                    ).count(),
                    'views_30_days': asset.auditlog_set.filter(
                        action='view',
                        timestamp__gte=thirty_days_ago
                    ).count(),
                    'transfers_total': asset.transfers.filter(
                        state='completed'
                    ).count(),
                    'maintenance_total': asset.maintenance_records.filter(
                        status='completed'
                    ).count(),
                }
            except Exception as e:
                context['utilization_stats'] = {
                    'scans_30_days': 0,
                    'views_30_days': 0,
                    'transfers_total': 0,
                    'maintenance_total': 0,
                    'error': True,
                }
        
        # Device and scan context
        context['device_info'] = self.get_device_info(
            self.request.META.get('HTTP_USER_AGENT', '')
        )
        context['is_mobile'] = 'Mobile' in context['device_info']
        context['accessed_by_uuid'] = True
        
        return context
    
    def get_status_badge_class(self, status):
        """Return Bootstrap badge class for asset status."""
        status_map = {
            'active': 'bg-success',
            'in_maintenance': 'bg-warning text-dark',
            'retired': 'bg-secondary',
            'lost': 'bg-danger',
            'deleted': 'bg-danger text-white',  # Changed from bg-dark for visibility
            'transferred': 'bg-info',
        }
        return status_map.get(status, 'bg-primary')

    def get(self, request, *args, **kwargs):
        asset = self.get_object()
        is_internal = request.GET.get('internal') == '1'
        user_agent = request.META.get('HTTP_USER_AGENT', '')
        device_info = self.get_device_info(user_agent)
        
        # Enhanced audit logging with more metadata
        ip_address = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
        if ',' in ip_address:
            ip_address = ip_address.split(',')[0].strip()
        
        metadata = {
            'device_info': device_info,
            'user_agent': user_agent[:200],
            'ip_address': ip_address,
            'is_authenticated': request.user.is_authenticated,
            'access_type': 'internal' if is_internal else 'qr_scan',
            'referer': request.META.get('HTTP_REFERER', '')[:200],
        }
        
        if not is_internal:
            # Log as scan (QR code access)
            log_audit(
                request.user if request.user.is_authenticated else None,
                'scan',
                asset,
                f'Asset scanned via QR code. Device: {device_info}. Auth: {request.user.is_authenticated}',
                metadata=metadata
            )
        else:
            # Log as regular view (internal navigation)
            log_audit(
                request.user if request.user.is_authenticated else None,
                'view',
                asset,
                f'Asset viewed via internal navigation. Device: {device_info}',
                metadata=metadata
            )
        
        return super().get(request, *args, **kwargs)
    
    def render_to_response(self, context, **response_kwargs):
        """
        WORLD-CLASS: Add cache-busting headers to ensure fresh data.
        Prevents browser from caching stale asset data after edits.
        """
        response = super().render_to_response(context, **response_kwargs)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

# WORLD-CLASS: Export Preview Endpoint
# Inspired by ServiceNow ITAM, IBM Maximo, SAP EAM export preview features
@require_POST
@login_required
def asset_export_preview(request):
    """
    Generate preview of export data before actual export.
    
    Features:
    - Shows first 50 rows for quick preview
    - Returns total count for user awareness
    - Respects all filters (category, status, search, etc.)
    - Company-scoped (multi-tenancy)
    - Fast response (<500ms for most queries)
    
    Request Body (JSON):
        {
            "format": "xlsx|csv|pdf",
            "category": "category_id",
            "status": "active|in_maintenance|retired",
            "search": "search_term",
            "selected_ids": "1,2,3" (optional - preview specific assets)
        }
    
    Response (JSON):
        {
            "success": true,
            "preview_rows": [{"id": 1, "category": "...", ...}, ...],
            "total_count": 150,
            "preview_count": 50,
            "format": "xlsx",
            "filters_applied": {"category": "Computers", "status": "active"},
            "columns": ["id", "category", "status", ...]
        }
    
    Performance: < 500ms for 10K assets
    Security: Auth required, company-scoped, audit logged
    """
    import json
    
    user = request.user
    company = getattr(request, 'company', None)
    
    # Validation
    if not company:
        return JsonResponse({'success': False, 'error': 'Company context required'}, status=403)
    if not can(user, 'export_data'):
        return JsonResponse({'success': False, 'error': 'Insufficient permissions'}, status=403)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    
    # Extract parameters
    format_type = data.get('format', 'xlsx')
    selected_ids = data.get('selected_ids', '')
    
    # Build queryset (same logic as export)
    assets = Asset.objects.for_company(company).select_related('category', 'branch', 'assigned_to')
    
    # Filter by selected IDs if provided
    if selected_ids:
        id_list = [int(pk) for pk in str(selected_ids).split(',') if pk.strip().isdigit()]
        assets = assets.filter(pk__in=id_list)
    else:
        # Apply filters
        category = data.get('category')
        status = data.get('status')
        location = data.get('location')
        search = data.get('search')
        branch = data.get('branch')
        
        if category:
            assets = assets.filter(category__id=category)
        if status:
            assets = assets.filter(status=status)
        if location:
            assets = assets.filter(dynamic_data__location__icontains=location)
        if branch:
            assets = assets.filter(branch__id=branch)
        if search:
            assets = assets.filter(
                Q(dynamic_data__name__icontains=search) |
                Q(dynamic_data__model__icontains=search) |
                Q(description__icontains=search)
            )
    
    # Get total count
    total_count = assets.count()
    
    # Get preview (first 50 rows)
    preview_limit = 50
    preview_assets = assets[:preview_limit]
    
    # Serialize preview data
    preview_rows = []
    for asset in preview_assets:
        row = {
            'id': asset.id,
            'uuid': str(asset.uuid),
            'category': asset.category.name if asset.category else '-',
            'branch': asset.branch.name if asset.branch else '-',
            'status': asset.get_status_display(),
            'assigned_to': asset.assigned_to.get_full_name() if asset.assigned_to else 'Unassigned',
            'created_at': asset.created_at.strftime('%Y-%m-%d %H:%M'),
        }
        
        # Add dynamic fields
        if asset.dynamic_data:
            for key, value in asset.dynamic_data.items():
                row[key] = value if value is not None else '-'
        
        preview_rows.append(row)
    
    # Determine columns
    if preview_rows:
        columns = list(preview_rows[0].keys())
    else:
        columns = ['id', 'category', 'branch', 'status', 'assigned_to', 'created_at']
    
    # Build filters_applied summary
    filters_applied = {}
    if data.get('category'):
        try:
            cat = AssetCategory.objects.for_company(company).get(pk=data['category'])
            filters_applied['category'] = cat.name
        except AssetCategory.DoesNotExist:
            pass
    if data.get('status'):
        filters_applied['status'] = data['status']
    if data.get('search'):
        filters_applied['search'] = data['search']
    if data.get('branch'):
        try:
            from tenancy.models import Branch
            br = Branch.objects.filter(company=company, id=data['branch']).first()
            if br:
                filters_applied['branch'] = br.name
        except:
            pass
    
    # Audit log
    from audit.utils import log_audit
    log_audit(
        user, 'export_preview', None,
        f'Export preview: {total_count} assets, format: {format_type}',
        company=company,
        metadata={'total_count': total_count, 'format': format_type, 'filters': filters_applied}
    )
    
    return JsonResponse({
        'success': True,
        'preview_rows': preview_rows,
        'total_count': total_count,
        'preview_count': min(preview_limit, total_count),
        'format': format_type,
        'filters_applied': filters_applied,
        'columns': columns,
        'has_more': total_count > preview_limit
    })


# Export endpoint (robust, supports GET and POST, individual/bulk)
def asset_export(request):
    # Authorization: require login and explicit export permission
    user = request.user
    if not user.is_authenticated:
        return HttpResponse('Authentication required', status=401)
    if not can(user, 'export_data'):
        return HttpResponse('Forbidden: insufficient permissions', status=403)
    company = getattr(request, 'company', None)
    if not company:
        return HttpResponse('Company context required', status=403)
    data_source = request.POST if request.method == 'POST' else request.GET
    format = data_source.get('format', 'csv')
    columns = data_source.getlist('columns') if hasattr(data_source, 'getlist') else data_source.get('columns', [])
    selected_ids = data_source.get('selected_ids', '')
    # Reuse AssetListView filtering logic
    assets = Asset.objects.for_company(company)
    # If selected_ids provided, filter by those
    if selected_ids:
        id_list = [int(pk) for pk in selected_ids.split(',') if pk.strip().isdigit()]
        assets = assets.filter(pk__in=id_list)
    else:
        category = data_source.get('category')
        status = data_source.get('status')
        location = data_source.get('location')
        search = data_source.get('search')
        if category:
            assets = assets.filter(category__id=category)
            fields = AssetCategoryField.objects.for_company(company).filter(category_id=category)
            for field in fields:
                val = data_source.get(f'dyn_{field.key}')
                if val:
                    if field.type == 'text':
                        assets = assets.filter(**{f'dynamic_data__{field.key}__icontains': val})
                    elif field.type == 'number':
                        try:
                            assets = assets.filter(**{f'dynamic_data__{field.key}': float(val)})
                        except ValueError:
                            pass
                    elif field.type == 'date':
                        assets = assets.filter(**{f'dynamic_data__{field.key}': val})
        if status:
            assets = assets.filter(status=status)
        if location:
            assets = assets.filter(dynamic_data__location__icontains=location)
        if search:
            assets = assets.filter(
                Q(dynamic_data__name__icontains=search) |
                Q(dynamic_data__model__icontains=search) |
                Q(description__icontains=search)
            )
    # Prepare data
    data = []
    for asset in assets:
        row = {}
        # Core fields
        row['ID'] = asset.pk
        row['Category'] = asset.category.name
        row['Status'] = asset.status
        row['Assigned To'] = str(asset.assigned_to) if asset.assigned_to else ''
        row['Created'] = asset.created_at.strftime('%Y-%m-%d %H:%M')
        row['Updated'] = asset.updated_at.strftime('%Y-%m-%d %H:%M')
        # Dynamic fields
        for k, v in asset.dynamic_data.items():
            row[k] = v
        data.append(row)
    # Filter columns
    if columns:
        data = [ {k: row.get(k, '') for k in columns} for row in data ]
    # Large export warning
    large_export = len(data) > 1000
    # Log export
    log = ExportLog.objects.create(
        user=request.user if request.user.is_authenticated else None,
        format=format,
        columns=columns,
        filters=data_source.dict() if hasattr(data_source, 'dict') else {},
        success=True
    )
    try:
        if format == 'csv':
            df = pd.DataFrame(data)
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = (f'attachment; filename="assets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"')
            if large_export:
                response['X-Export-Warning'] = 'Export is very large and may take time.'
            df.to_csv(response, index=False)
            log_audit(request.user, 'export', None, 'Assets exported as CSV')
            return response
        elif format == 'xlsx':
            df = pd.DataFrame(data)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = (f'attachment; filename="assets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"')
            if large_export:
                response['X-Export-Warning'] = 'Export is very large and may take time.'
            with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Assets')
            log_audit(request.user, 'export', None, 'Assets exported as Excel')
            return response
        elif format == 'pdf':
            all_columns = list(columns or (list(data[0].keys()) if data else []))
            core_default = ['ID', 'Category', 'Status', 'Assigned To', 'Created', 'Updated']
            core_columns = [c for c in all_columns if c in core_default]
            dynamic_attributes_map = {}
            for row in data:
                try:
                    aid = row.get('ID') or row.get('id')
                except Exception:
                    aid = None
                attrs = []
                for k, v in row.items():
                    if k not in core_columns:
                        attrs.append({'label': k, 'value': v})
                if aid is not None:
                    dynamic_attributes_map[aid] = attrs

            html_string = render_to_string('assets/export_pdf.html', {
                'assets': data,
                'columns': all_columns,
                'core_columns': core_columns,
                'dynamic_attributes_map': dynamic_attributes_map,
                'tenant_name': getattr(getattr(request, 'company', None), 'name', ''),
                'generated_by': (request.user.get_full_name() or request.user.username) if request.user.is_authenticated else 'System',
                'logo_url': settings.STATIC_URL + 'img/logo.png',
                'export_date': datetime.now(),
            })
            import os
            fd, temp_path = tempfile.mkstemp(suffix='.pdf')
            os.close(fd)
            try:
                HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(temp_path)
                with open(temp_path, 'rb') as f:
                    pdf = f.read()
            finally:
                os.remove(temp_path)
            response = HttpResponse(pdf, content_type='application/pdf')
            response["Content-Disposition"] = f'attachment; filename="assets_{timestamp}.pdf"'
            if large_export:
                response['X-Export-Warning'] = 'Export is very large and may take time.'
            log_audit(request.user, 'export', None, 'Assets exported as PDF')
            return response
        else:
            log.success = False
            log.error_message = 'Invalid export format'
            log.save()
            return HttpResponse('Invalid export format', status=400)
    except Exception as e:
        log.success = False
        log.error_message = str(e)
        log.save()
        return HttpResponse(f'Export failed: {e}', status=500)

@login_required
@user_passes_test(is_admin_or_manager, login_url='users:login')
def download_import_template(request):
    # Enforce matrix: require create permission
    if not can(request.user, 'create_assets'):
        return HttpResponse('Forbidden: insufficient permissions', status=403)
    # Generate Excel template for selected category
    category_id = request.GET.get('category')
    if not category_id:
        return HttpResponse('Category required', status=400)
    company = getattr(request, 'company', None)
    if not company:
        return HttpResponse('Company context required', status=403)
    category = AssetCategory.objects.for_company(company).get(pk=category_id)
    dynamic_fields = AssetCategoryField.objects.for_company(company).filter(category=category)
    wb = openpyxl.Workbook()
    ws = wb.active
    # Core fields
    columns = ['status', 'description', 'assigned_to']
    # Dynamic fields
    columns += [f.key for f in dynamic_fields]
    ws.append(columns)
    # Add sample row
    if request.GET.get('example') == '1':
        # Provide realistic example data
        sample = ['active', 'Laptop for CEO', 'manager']
        for f in dynamic_fields:
            if f.key == 'serial_number':
                sample.append('SN123456')
            elif f.key == 'location':
                sample.append('HQ Office')
            elif f.key == 'purchase_date':
                sample.append('2023-01-15')
            else:
                sample.append('Example')
        ws.append(sample)
    else:
        sample = ['active', 'Sample asset', 'manager'] + ['' for _ in dynamic_fields]
        ws.append(sample)
    # Style header
    for i, col in enumerate(columns, 1):
        ws[f'{get_column_letter(i)}1'].font = openpyxl.styles.Font(bold=True)
    # Return as response
    import os
    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    try:
        wb.save(temp_path)
        with open(temp_path, 'rb') as f:
            file_data = f.read()
    finally:
        os.remove(temp_path)
    response = HttpResponse(file_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if request.GET.get('example') == '1':
        response['Content-Disposition'] = (f'attachment; filename=asset_import_example_{category.name}.xlsx')
    else:
        response['Content-Disposition'] = (f'attachment; filename=asset_import_template_{category.name}.xlsx')
    return response

@method_decorator(user_passes_test(is_admin_or_manager, login_url='users:login'), name='dispatch')
class AssetBulkImportView(View):
    template_name = 'assets/asset_bulk_import.html'

    def get(self, request):
        # Enforce matrix: require create permission
        if not can(request.user, 'create_assets'):
            return HttpResponse('Forbidden: insufficient permissions', status=403)
        # Step 1: Select category, download template
        company = getattr(self.request, 'company', None)
        categories = AssetCategory.objects.for_company(company)
        selected_category = request.GET.get('category')
        step = request.GET.get('step', '1')
        context = {'categories': categories, 'selected_category': selected_category, 'step': step}
        return render(request, self.template_name, context)

    def post(self, request):
        # Enforce matrix: require create permission
        if not can(request.user, 'create_assets'):
            return HttpResponse('Forbidden: insufficient permissions', status=403)
        # Step 2: Upload and preview file
        company = getattr(self.request, 'company', None)
        if not company:
            messages.error(request, 'Company context required for bulk import.')
            return HttpResponse('Company context required', status=403)
        
        branch = getattr(self.request, 'branch', None)
        categories = AssetCategory.objects.for_company(company)
        selected_category = request.POST.get('category')
        if selected_category:
            categories = categories.filter(pk=selected_category)
        step = request.POST.get('step', '2')
        file = request.FILES.get('import_file')
        preview_data = []
        errors = []
        columns = []
        import_file = request.POST.get('import_file')
        
        if step == '3':
            # Final confirmation: import assets with QR code generation
            # Re-read the file from temp storage
            file_path = default_storage.path(import_file)
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(columns, row))
                    preview_data.append(row_data)
            except Exception as e:
                errors.append(f'Failed to parse file: {e}')
                messages.error(request, f'Failed to parse file: {e}')
                return render(request, self.template_name, {
                    'categories': categories,
                    'selected_category': selected_category,
                    'step': '1',
                    'errors': errors
                })
            
            dynamic_fields = AssetCategoryField.objects.for_company(company).filter(category_id=selected_category)
            success_count = 0
            fail_count = 0
            fail_rows = []
            imported_assets = []
            qr_generation_failures = []
            
            # Process each row individually with proper error handling
            for i, row in enumerate(preview_data):
                try:
                    # Validate required fields
                    for field in dynamic_fields:
                        if field.required and not row.get(field.key):
                            raise ValueError(f'Missing required field {field.label}')
                    
                    with transaction.atomic():
                        # Create asset with company and branch context
                        asset = Asset(
                            company=company,
                            branch=branch,
                            category_id=selected_category,
                            status=row.get('status', 'active'),
                            description=row.get('description', ''),
                        )
                        
                        # Assign user if provided
                        assigned_to = row.get('assigned_to')
                        if assigned_to:
                            from users.models import User
                            user_obj = User.objects.filter(username=assigned_to, company=company).first()
                            if user_obj:
                                asset.assigned_to = user_obj
                            else:
                                # Log warning but continue
                                print(f"Warning: User '{assigned_to}' not found in company {company.name}")
                        
                        # Dynamic fields
                        dyn_data = {}
                        for field in dynamic_fields:
                            value = row.get(field.key)
                            # Convert datetime objects to ISO format strings
                            if isinstance(value, datetime):
                                value = value.isoformat()
                            # Handle date objects from Excel
                            elif hasattr(value, 'date') and callable(getattr(value, 'date')):
                                value = value.date().isoformat()
                            # Handle empty values
                            elif value is None or value == '':
                                value = None
                            dyn_data[field.key] = value
                        asset.dynamic_data = dyn_data
                        
                        # Initial save to generate UUID and establish DB record
                        asset.save()
                        
                        # Generate QR code with enhanced error handling
                        qr_generated = False
                        try:
                            import os
                            from django.conf import settings
                            
                            # Ensure QR codes directory exists
                            qr_dir = os.path.join(settings.MEDIA_ROOT, 'qr_codes')
                            os.makedirs(qr_dir, exist_ok=True)
                            
                            # Build absolute URL for QR code
                            base_url = request.build_absolute_uri('/')[:-1]
                            qr_url = f"{base_url}/assets/{asset.uuid}/"
                            
                            # Generate QR code with high quality settings
                            qr = qrcode.QRCode(
                                version=1,
                                error_correction=qrcode.constants.ERROR_CORRECT_H,
                                box_size=10,
                                border=4,
                            )
                            qr.add_data(qr_url)
                            qr.make(fit=True)
                            
                            img = qr.make_image(fill_color="black", back_color="white")
                            buffer = BytesIO()
                            img.save(buffer, 'PNG')
                            buffer.seek(0)
                            
                            # Save QR code to asset
                            asset.qr_code.save(
                                f"asset_{asset.uuid}.png",
                                ContentFile(buffer.getvalue()),
                                save=False
                            )
                            qr_generated = True
                            
                        except Exception as qr_error:
                            qr_generation_failures.append({
                                'row': i+2,
                                'asset_id': asset.pk,
                                'error': str(qr_error)
                            })
                            print(f"QR code generation failed for bulk import row {i+2}: {qr_error}")
                            # Continue without QR code - asset is still valid
                        
                        # Final save with QR code (if generated)
                        asset.save()
                        
                        # Track successful import
                        imported_assets.append(asset)
                        
                        # Log audit events after successful save
                        if asset.assigned_to:
                            log_audit(
                                request.user,
                                ASSIGN_ACTION,
                                asset,
                                f'Asset assigned to {asset.assigned_to.username} via bulk import (row {i+2})',
                                related_user=asset.assigned_to
                            )
                        
                        log_audit(
                            request.user,
                            'create',
                            asset,
                            f'Asset imported via bulk import (row {i+2}) - QR code: {"generated" if qr_generated else "failed"}'
                        )
                        success_count += 1
                        
                except Exception as e:
                    fail_count += 1
                    fail_rows.append({'row': i+2, 'error': str(e)})
                    print(f"Failed to import row {i+2}: {e}")
            
            # Log bulk import summary
            try:
                from assets.models import ExportLog
                ExportLog.objects.create(
                    user=request.user,
                    format='import',
                    columns=columns,
                    filters={'category': selected_category, 'company': company.name},
                    success=(fail_count == 0),
                    error_message='; '.join([f"Row {r['row']}: {r['error']}" for r in fail_rows]) if fail_rows else ''
                )
            except Exception as log_error:
                print(f"Failed to create export log: {log_error}")
            
            # Log comprehensive audit for bulk import operation
            log_audit(
                request.user,
                'bulk_import',
                None,
                f'Bulk import completed: {success_count} assets imported, {fail_count} failed, {len(qr_generation_failures)} QR generation failures',
                metadata={
                    'success_count': success_count,
                    'fail_count': fail_count,
                    'qr_failures': len(qr_generation_failures),
                    'category': selected_category,
                    'company': company.name
                }
            )
            
            # Prepare success message
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} asset(s) with QR codes.')
            if qr_generation_failures:
                messages.warning(
                    request,
                    f'{len(qr_generation_failures)} asset(s) imported but QR code generation failed. '
                    'You can regenerate QR codes from the asset detail page.'
                )
            if fail_count > 0:
                messages.error(request, f'{fail_count} row(s) failed to import. See details below.')
            
            context = {
                'categories': categories,
                'selected_category': selected_category,
                'step': 'done',
                'success_count': success_count,
                'fail_count': fail_count,
                'fail_rows': fail_rows,
                'qr_failures': qr_generation_failures,
                'imported_assets': imported_assets,
            }
            return render(request, self.template_name, context)
        if not file or not selected_category:
            messages.error(request, 'Please select a category and upload a file.')
            return render(request, self.template_name, {'categories': categories, 'selected_category': selected_category, 'step': '1'})
        # Save file temporarily
        tmp_path = default_storage.save('tmp/' + file.name, file)
        file_path = default_storage.path(tmp_path)
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(columns, row))
                preview_data.append(row_data)
        except Exception as e:
            errors.append(f'Failed to parse file: {e}')
        # Validate rows (basic, more in confirm step)
        dynamic_fields = AssetCategoryField.objects.for_company(company).filter(category_id=selected_category)
        for i, row in enumerate(preview_data):
            for field in dynamic_fields:
                if field.required and not row.get(field.key):
                    errors.append(f'Row {i+2}: Missing required field {field.label}')
        context = {
            'categories': categories,
            'selected_category': selected_category,
            'step': step,
            'preview_data': preview_data,
            'columns': columns,
            'errors': errors,
            'import_file': tmp_path,
        }
        return render(request, self.template_name, context)

    def put(self, request):
        # Step 3: Confirm import (AJAX or form submit)
        # Not implemented yet
        pass

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        context['role'] = getattr(user, 'role', 'user')
        return context

@login_required
@require_GET
def dashboard_summary_api(request):
    """WORLD-CLASS: Dashboard summary with multi-tenancy, RBAC, and performance optimization.
    
    CRITICAL FIX: Uses unified AssetFilteringService to ensure 100% consistency
    with asset list view. This eliminates data discrepancies.
    """
    from datetime import timedelta
    from django.utils import timezone
    from django.core.cache import cache
    from django.db.models import Count, Q
    from assets.services.filtering import asset_filtering_service
    
    user = request.user
    role = getattr(user, 'role', 'user')
    company = getattr(request, 'company', None)
    active_branch = getattr(request, 'branch', None)  # Active branch from session

    # SECURITY: Enforce company requirement
    if company is None:
        return JsonResponse(
            {
                'error': 'Company context required',
                'kpis': {},
                'by_category': {},
                'trends': {},
                'role': role,
                'user_id': user.id,
            },
            status=403,
        )

    # PERFORMANCE: Cache key for user-specific dashboard data (5 min TTL)
    # WORLD-CLASS: Include branch in cache key for branch-aware caching
    branch_suffix = f"_branch_{active_branch.id}" if active_branch else "_all_branches"
    cache_key = f'dashboard_summary_{company.id}_{role}_{user.id}{branch_suffix}'
    try:
        cached_data = cache.get(cache_key)
        if cached_data:
            return JsonResponse(cached_data)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Cache get failed for dashboard summary: {e}")

    # WORLD-CLASS FIX: Use unified filtering service for 100% consistency
    # This ensures dashboard metrics match asset list counts EXACTLY
    qs = asset_filtering_service.get_base_queryset(company, user, request)
    
    # Add select_related for performance and defer heavy fields
    qs = qs.select_related('category', 'branch', 'assigned_to').defer('dynamic_data', 'qr_code')
    
    # CRITICAL FIX: Apply branch filter if active branch is set (branch switching support)
    if active_branch:
        qs = qs.filter(branch=active_branch)
    
    # PERFORMANCE: Use aggregate queries instead of multiple count() calls
    status_counts = qs.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(status=Asset.STATUS_ACTIVE)),
        maintenance=Count('id', filter=Q(status=Asset.STATUS_IN_MAINTENANCE)),
        retired=Count('id', filter=Q(status=Asset.STATUS_RETIRED)),
        lost=Count('id', filter=Q(status=Asset.STATUS_LOST)),
        assigned=Count('id', filter=Q(assigned_to__isnull=False)),
        unassigned=Count('id', filter=Q(assigned_to__isnull=True)),
        transferred=Count('id', filter=Q(status=Asset.STATUS_TRANSFERRED)),
    )
    
    total_assets = status_counts['total']
    active_assets = status_counts['active']
    maintenance_assets = status_counts['maintenance']
    retired_assets = status_counts['retired']
    lost_assets = status_counts['lost']
    assigned_assets = status_counts['assigned']
    unassigned_assets = status_counts['unassigned']
    transferred_assets = status_counts['transferred']
    
    # WORLD-CLASS: Warranty expiry with proper date handling
    warranty_expiring_soon = 0
    try:
        soon = timezone.now() + timedelta(days=30)
        now_date = timezone.now().date().isoformat()
        soon_date = soon.date().isoformat()
        
        # Simplified warranty check (avoid complex JSON queries that can cause locks)
        # Only check if we have a reasonable number of assets
        if total_assets < 1000:
            warranty_expiring_soon = qs.filter(
                Q(dynamic_data__warranty_expiry__lte=soon_date, dynamic_data__warranty_expiry__gte=now_date) |
                Q(dynamic_data__warranty_end__lte=soon_date, dynamic_data__warranty_end__gte=now_date) |
                Q(dynamic_data__warranty_expiration__lte=soon_date, dynamic_data__warranty_expiration__gte=now_date)
            ).count()
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Warranty expiry calculation failed: {e}")
        warranty_expiring_soon = 0
    
    # WORLD-CLASS: Users with no assets (admin/manager only)
    users_with_no_assets = 0
    if role in ('admin', 'manager'):
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            # Simplified calculation to avoid subquery locks
            total_users = User.objects.filter(company=company, is_active=True).count()
            users_with_assets_count = qs.filter(assigned_to__isnull=False).values('assigned_to_id').distinct().count()
            users_with_no_assets = max(0, total_users - users_with_assets_count)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Users with no assets calculation failed: {e}")
            users_with_no_assets = 0
    
    # WORLD-CLASS: Pending approvals (transfer approvals) - RBAC-aware
    approvals_pending = 0
    try:
        from .models import AssetTransfer
        
        if role == 'user':
            # Users see transfers where they are receiver (pending their approval)
            approvals_pending = AssetTransfer.objects.filter(
                company=company,
                to_user=user,
                state=AssetTransfer.TransferState.PENDING_RECEIVER
            ).count()
        elif role == 'manager':
            # Managers see transfers in their accessible branches
            from tenancy.policy_service import PolicyService
            try:
                accessible_branch_ids = PolicyService.get_accessible_branches(user, company)
                if accessible_branch_ids:
                    approvals_pending = AssetTransfer.objects.filter(
                        company=company,
                        state__in=[
                            AssetTransfer.TransferState.PENDING_RECEIVER,
                            AssetTransfer.TransferState.RECEIVER_APPROVED,
                            AssetTransfer.TransferState.AWAITING_ADMIN
                        ],
                        asset__branch_id__in=accessible_branch_ids
                    ).count()
                else:
                    approvals_pending = 0
            except Exception:
                # Fallback: show transfers where manager is involved
                approvals_pending = AssetTransfer.objects.filter(
                    company=company,
                    state__in=[
                        AssetTransfer.TransferState.PENDING_RECEIVER,
                        AssetTransfer.TransferState.RECEIVER_APPROVED
                    ]
                ).count()
        elif role == 'admin':
            # Admins see all pending transfers in company
            approvals_pending = AssetTransfer.objects.filter(
                company=company,
                state__in=[
                    AssetTransfer.TransferState.PENDING_RECEIVER,
                    AssetTransfer.TransferState.RECEIVER_APPROVED,
                    AssetTransfer.TransferState.AWAITING_ADMIN
                ]
            ).count()
    except (ImportError, AttributeError) as e:
        approvals_pending = 0
    
    # PERFORMANCE: Category breakdown with single query
    category_counts = (
        qs.values('category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]  # Top 10 categories
    )
    by_category = {entry['category__name'] or 'Uncategorized': entry['count'] for entry in category_counts}
    
    # WORLD-CLASS: Trends with proper calculation
    month_ago = timezone.now() - timedelta(days=30)
    total_assets_month_ago = qs.filter(created_at__lte=month_ago).count()
    
    if total_assets_month_ago > 0:
        change = ((total_assets - total_assets_month_ago) / total_assets_month_ago) * 100
        total_assets_monthly_change = f"{change:+.1f}%"  # Include + sign for positive
    elif total_assets > 0:
        total_assets_monthly_change = "+100.0%"  # All new assets
    else:
        total_assets_monthly_change = "0.0%"
    
    # WORLD-CLASS: Additional metrics for comprehensive dashboard
    needs_repair = maintenance_assets  # Alias for clarity
    
    response_data = {
        'kpis': {
            'total_assets': total_assets,
            'active_assets': active_assets,
            'maintenance_assets': maintenance_assets,
            'needs_repair': needs_repair,  # Alias for frontend compatibility
            'retired_assets': retired_assets,
            'lost_assets': lost_assets,
            'assigned_assets': assigned_assets,
            'unassigned_assets': unassigned_assets,
            'warranty_expiring_soon': warranty_expiring_soon,
            'transferred_assets': transferred_assets,
            'users_with_no_assets': users_with_no_assets,
            'approvals_pending': approvals_pending,
        },
        'by_category': by_category,
        'trends': {
            'total_assets_monthly_change': total_assets_monthly_change,
        },
        'role': role,
        'user_id': user.id,
        'company_id': company.id,
        'timestamp': timezone.now().isoformat(),
    }
    
    # PERFORMANCE: Cache for 5 minutes with error handling
    try:
        cache.set(cache_key, response_data, 300)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Cache set failed for dashboard summary: {e}")
    
    return JsonResponse(response_data)

@login_required
@require_GET
def dashboard_activity_api(request):
    """
    WORLD-CLASS: Dashboard activity API with proper role-based filtering.
    - Admin: See all company activities
    - Manager: See only activities from their assigned branches
    - User: See only their own activities
    """
    user = request.user
    role = getattr(user, 'role', 'user')
    company = getattr(request, 'company', None)
    active_branch = getattr(request, 'branch', None)  # Active branch from session
    
    # Base query: company-scoped
    logs = AuditLog.objects.filter(company=company).order_by('-timestamp')
    
    # WORLD-CLASS: Role-based filtering with security-first approach
    if role == 'user':
        # Regular users: Only see their own activities
        logs = logs.filter(user=user)
    elif role == 'manager':
        # Managers: Only see activities from their assigned branches
        from tenancy.models import UserBranch
        
        # Get manager's assigned branches
        user_branch_ids = UserBranch.objects.filter(
            user=user
        ).values_list('branch_id', flat=True)
        
        if user_branch_ids:
            # Filter logs by branch OR logs related to assets in those branches
            from django.db.models import Q
            logs = logs.filter(
                Q(branch_id__in=user_branch_ids) |
                Q(asset__branch_id__in=user_branch_ids)
            )
        else:
            # SECURITY: If manager has no branch assignments, show nothing
            logs = logs.none()
    # else: role == 'admin' → No additional filtering, admins see all company logs
    
    # Apply active branch filter if user has selected a specific branch
    if active_branch and role != 'admin':
        logs = logs.filter(
            Q(branch=active_branch) | 
            Q(asset__branch=active_branch)
        )
    
    # Limit to recent 20 activities
    logs = logs[:20]
    data = [
        {
            'user': str(log.user) if log.user else '',
            'action': log.action,
            'asset': str(log.asset) if log.asset else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        }
        for log in logs
    ]
    # TODO: Consider logging dashboard activity API access for auditability
    return JsonResponse({'activity': data})

@login_required
@require_GET
def notifications_api(request):
    """
    Minimal read-only Notifications API (Option A)
    - Returns recent AuditLog-derived notifications for the current user context
    - No unread/read state (unread_count is reported as 0)
    Query params:
      - limit: max number of items to return (default 5, max 20)
    """
    user = request.user
    role = getattr(user, 'role', 'user')
    try:
        limit = int(request.GET.get('limit', 5))
    except (TypeError, ValueError):
        limit = 5
    limit = max(1, min(limit, 20))

    # Base queryset: latest audit logs
    logs = AuditLog.objects.all().order_by('-timestamp')

    # Role-based visibility
    # Only admins see all. Managers and regular users are restricted.
    if role != 'admin':
        # Users and managers: only their own actions or actions on assets assigned to them
        logs = logs.filter(
            Q(user=user) | Q(asset__assigned_to=user)
        )

    logs = logs[:limit]

    def icon_for_action(action: str) -> str:
        mapping = {
            'create': 'bi-plus-circle',
            'edit': 'bi-pencil-square',
            'assign': 'bi-person-check',
            'unassign': 'bi-person-dash',
            'status_change': 'bi-arrow-left-right',
            'maintenance': 'bi-tools',
            'retire': 'bi-archive',
            'lost': 'bi-exclamation-triangle',
            'scan': 'bi-qr-code-scan',
            'export': 'bi-filetype-csv',
            'warranty_change': 'bi-shield-check',
            'transfer': 'bi-arrow-repeat',
            'view': 'bi-eye',
        }
        return mapping.get((action or '').lower(), 'bi-bell')

    items = []
    for log in logs:
        title = ''
        # Build a concise title from action and asset
        action_label = (log.action or 'Activity').replace('_', ' ').title()
        asset_label = str(log.asset) if getattr(log, 'asset', None) else ''
        if asset_label:
            title = f"{action_label}: {asset_label}"
        else:
            title = action_label

        # Attempt to build a deep link to the asset if available
        url = ''
        try:
            if log.asset and getattr(log.asset, 'uuid', None):
                url = request.build_absolute_uri(f"/assets/{log.asset.uuid}/?internal=1")
        except Exception:
            url = ''

        items.append({
            'title': title,
            'message': log.details or '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'icon': icon_for_action(log.action),
            'url': url,
        })

    return JsonResponse({
        'unread_count': 0,  # No unread tracking in Option A
        'items': items,
        'limit': limit,
    })

@login_required
@require_GET
def dashboard_scan_logs_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action='view').order_by('-timestamp')
    if role == 'user':
        logs = logs.filter(user=user)
    logs = logs[:5]
    data = [
        {
            'user': str(log.user) if log.user else '',
            'asset': str(log.asset) if log.asset else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
        }
        for log in logs
    ]
    # TODO: Consider logging dashboard scan logs API access for auditability
    return JsonResponse({'scan_logs': data})

@login_required
@require_GET
def dashboard_chart_data_api(request):
    """WORLD-CLASS: Chart data API with multi-tenancy, RBAC, and performance optimization."""
    from django.db.models import Count
    from django.utils import timezone
    from django.core.cache import cache
    import calendar
    import datetime
    
    user = request.user
    role = getattr(user, 'role', 'user')
    company = getattr(request, 'company', None)
    branch = getattr(request, 'branch', None)
    
    # SECURITY: Enforce company requirement
    if company is None:
        return JsonResponse({'error': 'Company context required'}, status=403)
    
    chart = request.GET.get('chart')
    if not chart or chart not in {'category', 'acquisition', 'department', 'location'}:
        return JsonResponse({'error': 'Invalid or missing chart type'}, status=400)
    
    # PERFORMANCE: Cache key for chart data (10 min TTL)
    cache_key = f'dashboard_chart_{company.id}_{role}_{user.id}_{chart}'
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data)
    
    # WORLD-CLASS FIX: Use unified filtering service for consistency
    from assets.services.filtering import asset_filtering_service
    qs = asset_filtering_service.get_base_queryset(company, user, request)
    
    # Add select_related for performance
    qs = qs.select_related('category', 'branch')
    data = []
    labels = []
    # 1. Assets by Category
    if chart == 'category':
        agg = qs.values('category__name').annotate(count=Count('id')).order_by('-count')
        labels = [a['category__name'] for a in agg]
        data = [a['count'] for a in agg]
        response_data = {'chart': 'assets_by_category', 'labels': labels, 'data': data, 'role': role}
        cache.set(cache_key, response_data, 600)  # Cache for 10 minutes
        return JsonResponse(response_data)
    # 2. Asset Acquisition Over Time (last 12 months)
    elif chart == 'acquisition':
        now = timezone.now()
        months = [(now - datetime.timedelta(days=30*i)).strftime('%Y-%m') for i in reversed(range(12))]
        month_counts = {m: 0 for m in months}
        for asset in qs:
            m = asset.created_at.strftime('%Y-%m')
            if m in month_counts:
                month_counts[m] += 1
        labels = list(month_counts.keys())
        data = list(month_counts.values())
        response_data = {'chart': 'acquisition_over_time', 'labels': labels, 'data': data, 'role': role}
        cache.set(cache_key, response_data, 600)
        return JsonResponse(response_data)
    # 3. Assets by Department (dynamic field)
    elif chart == 'department':
        from django.db.models import Count, Value as V
        from django.db.models.functions import Coalesce
        from django.db.models.expressions import RawSQL
        try:
            # Prefer ORM aggregation with coalesced keys for performance
            from django.db.models.functions import Cast, Coalesce
            from django.db.models import CharField, Value as V
            qs_with_dept = qs.annotate(
                department=Coalesce(
                    Cast('dynamic_data__department', CharField()),
                    Cast('dynamic_data__Department', CharField()),
                    Cast('dynamic_data__dept', CharField()),
                    Cast('dynamic_data__Dept', CharField()),
                    Cast('dynamic_data__assigned_department', CharField()),
                    V('Unspecified')
                )
            )
            agg = qs_with_dept.values('department').annotate(count=Count('id')).order_by('-count')
            labels = [a['department'] or 'Unspecified' for a in agg]
            data = [a['count'] for a in agg]
        except Exception:
            # Fallback to Python loop if ORM fails
            dept_counts = {}
            keys = ['department', 'Department', 'dept', 'Dept', 'assigned_department']
            for asset in qs:
                # Try multiple keys and normalize value
                raw = None
                for k in keys:
                    if k in asset.dynamic_data:
                        raw = asset.dynamic_data.get(k)
                        break
                dept = (str(raw).strip() if raw else '') or 'Unspecified'
                dept_counts[dept] = dept_counts.get(dept, 0) + 1
            labels = list(dept_counts.keys())
            data = list(dept_counts.values())
        response_data = {'chart': 'assets_by_department', 'labels': labels, 'data': data, 'role': role}
        cache.set(cache_key, response_data, 600)
        return JsonResponse(response_data)
    # 4. Assets by Location (dynamic field)
    elif chart == 'location':
        from django.db.models import Count, Value as V
        from django.db.models.functions import Coalesce
        from django.db.models.expressions import RawSQL
        try:
            from django.db.models.functions import Cast, Coalesce
            from django.db.models import CharField, Value as V
            qs_with_loc = qs.annotate(
                location=Coalesce(
                    Cast('dynamic_data__location', CharField()),
                    Cast('dynamic_data__Location', CharField()),
                    Cast('dynamic_data__site', CharField()),
                    Cast('dynamic_data__Site', CharField()),
                    Cast('dynamic_data__office', CharField()),
                    Cast('dynamic_data__Office', CharField()),
                    Cast('dynamic_data__branch', CharField()),
                    Cast('dynamic_data__Branch', CharField()),
                    V('Unspecified')
                )
            )
            agg = qs_with_loc.values('location').annotate(count=Count('id')).order_by('-count')
            labels = [a['location'] or 'Unspecified' for a in agg]
            data = [a['count'] for a in agg]
        except Exception:
            loc_counts = {}
            keys = ['location', 'Location', 'site', 'Site', 'office', 'Office', 'branch', 'Branch']
            for asset in qs:
                # Try multiple keys and normalize value
                raw = None
                for k in keys:
                    if k in asset.dynamic_data:
                        raw = asset.dynamic_data.get(k)
                        break
                loc = (str(raw).strip() if raw else '') or 'Unspecified'
                loc_counts[loc] = loc_counts.get(loc, 0) + 1
            labels = list(loc_counts.keys())
            data = list(loc_counts.values())
        response_data = {'chart': 'assets_by_location', 'labels': labels, 'data': data, 'role': role}
        cache.set(cache_key, response_data, 600)
        return JsonResponse(response_data)
    
    # Should not reach here due to earlier validation
    return JsonResponse({'error': 'Invalid chart type'}, status=400)

def scope_audit_logs(logs, request, role):
    """Restrict audit logs to the current company and branch visibility."""
    company = getattr(request, 'company', None)
    user = getattr(request, 'user', None)
    if company:
        logs = logs.filter(company=company)

    # Superusers and admins: full company visibility
    if getattr(request.user, 'is_superuser', False) or role == 'admin':
        return logs

    # Active branch selected: for non-admins, restrict strictly to that branch or assets in that branch
    active_branch = getattr(request, 'branch', None)
    if active_branch is not None:
        # Restrict to selected branch or assets in that branch; include only the current user's global actions
        return logs.filter(
            Q(branch=active_branch) |
            Q(asset__branch=active_branch) |
            Q(branch__isnull=True, user=user)
        )

    # Manager role: restrict to assigned branches, plus their own global actions
    if role == 'manager':
        memberships = getattr(request, 'available_branches', [])
        branch_ids = [m.branch_id for m in memberships if getattr(m, 'branch_id', None)]
        if branch_ids:
            branch_scope = Q(branch_id__in=branch_ids) | Q(asset__branch_id__in=branch_ids)
            # Include only the manager's own global events (e.g., login/logout) but exclude others (e.g., admin)
            own_global = Q(branch__isnull=True, user=user)
            return logs.filter(branch_scope | own_global)
        # No branch assignments → show only the manager's own global actions
        return logs.filter(branch__isnull=True, user=user)

    # Standard users: callers will apply user-specific filters; as a safe default, only allow their own global actions
    return logs.filter(branch__isnull=True, user=user)


def paginate_logs(logs, request, default_size=10, max_size=50):
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', default_size))
        page_size = min(max(page_size, 1), max_size)
    except Exception:
        page = 1
        page_size = default_size
    paginator = Paginator(logs, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    return page_obj, paginator

@login_required
@require_GET
def recent_added_assets_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action__in=['create', 'add']).select_related('asset').order_by('-timestamp')
    logs = scope_audit_logs(logs, request, role)
    if role == 'user':
        logs = logs.filter(user=user)
    page_obj, paginator = paginate_logs(logs, request)
    data = [
        {
            'asset_id': log.asset.id if log.asset else None,
            'asset_uuid': str(log.asset.uuid) if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'user': str(log.user) if log.user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        }
        for log in page_obj
    ]
    return JsonResponse({'recent_added_assets': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
@require_GET
def recent_scans_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action='scan').select_related('asset').order_by('-timestamp')
    logs = scope_audit_logs(logs, request, role)
    if role == 'user':
        logs = logs.filter(user=user)
    page_obj, paginator = paginate_logs(logs, request)
    data = [
        {
            'asset_id': log.asset.id if log.asset else None,
            'asset_uuid': str(log.asset.uuid) if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'user': str(log.user) if log.user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
            'device_info': (log.metadata.get('device_info') if log.metadata else None) or 'Unknown',
        }
        for log in page_obj
    ]
    return JsonResponse({'recent_scans': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
@require_GET
def recent_transfers_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action='assign').select_related('asset').order_by('-timestamp')
    logs = scope_audit_logs(logs, request, role)
    if role == 'user':
        logs = logs.filter(user=user)
    page_obj, paginator = paginate_logs(logs, request)
    data = [
        {
            'asset_id': log.asset.id if log.asset else None,
            'asset_uuid': str(log.asset.uuid) if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'from_user': str(log.user) if log.user else '',
            'to_user': str(log.related_user) if log.related_user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        }
        for log in page_obj
    ]
    return JsonResponse({'recent_transfers': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
@require_GET
def recent_maintenance_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action='maintenance').select_related('asset').order_by('-timestamp')
    logs = scope_audit_logs(logs, request, role)
    if role == 'user':
        logs = logs.filter(user=user)
    page_obj, paginator = paginate_logs(logs, request)
    data = [
        {
            'asset_id': log.asset.id if log.asset else None,
            'asset_uuid': str(log.asset.uuid) if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'user': str(log.user) if log.user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        }
        for log in page_obj
    ]
    return JsonResponse({'recent_maintenance': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
@require_GET
def full_audit_log_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.all().select_related('asset').order_by('-timestamp')
    logs = scope_audit_logs(logs, request, role)
    if role == 'user':
        logs = logs.filter(user=user)
    action = request.GET.get('action')
    if action:
        logs = logs.filter(action=action)
    asset_id = request.GET.get('asset_id')
    if asset_id:
        logs = logs.filter(asset__id=asset_id)
    user_id = request.GET.get('user_id')
    if user_id:
        logs = logs.filter(user__id=user_id)
    page_obj, paginator = paginate_logs(logs, request, default_size=20, max_size=100)
    data = [
        {
            'id': log.id,
            'action': log.action,
            'asset_id': log.asset.id if log.asset else None,
            'asset_uuid': str(log.asset.uuid) if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'user': str(log.user) if log.user else '',
            'related_user': str(log.related_user) if log.related_user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
            'metadata': log.metadata,
        }
        for log in page_obj
    ]
    return JsonResponse({'audit_log': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
def user_assets_api(request):
    user = request.user
    assets = Asset.objects.filter(assigned_to=user)
    data = []
    for asset in assets:
        data.append({
            'name': str(asset),
            'serial': asset.dynamic_data.get('serial_number', ''),
            'assigned': str(asset.assigned_to) if asset.assigned_to else '',
            'status': asset.status,
        })
    return JsonResponse({'assets': data})

@login_required
def user_activity_api(request):
    user = request.user
    logs = AuditLog.objects.filter(user=user).order_by('-timestamp')[:20]
    data = []
    for log in logs:
        data.append({
            'action': log.action,
            'asset': str(log.asset) if log.asset else '',
            'time': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        })
    return JsonResponse({'logs': data})

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
@require_POST
def api_create_category(request):
    import json
    from assets.models import AssetCategory
    company = getattr(request, 'company', None)
    if not company:
        return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
    
    # Handle both JSON and form data
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            description = data.get('description', '').strip()
        else:
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data.'}, status=400)
    
    if not name:
        return JsonResponse({'success': False, 'error': 'Category name is required.'}, status=400)
    if AssetCategory.objects.for_company(company).filter(name__iexact=name).exists():
        return JsonResponse({'success': False, 'error': 'A category with this name already exists.'}, status=400)
    category = AssetCategory.objects.create(
        company=company, 
        name=name, 
        description=description,
        dynamic_fields={}
    )
    from audit.utils import log_audit
    log_audit(request.user, 'create', None, f'Category created: {name}')
    return JsonResponse({
        'success': True, 
        'category_id': category.id, 
        'category': {
            'id': category.id, 
            'name': category.name,
            'description': category.description
        }
    })

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
def api_categories(request):
    """Enhanced API endpoint with analytics for category management."""
    company = getattr(request, 'company', None)
    if not company:
        return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
    
    # Get categories with asset counts and resolve field counts from both relational and legacy JSON sources
    categories_qs = (
        AssetCategory.objects.for_company(company)
        .annotate(asset_count=Count('assets', distinct=True))
        .prefetch_related('fields')
        .order_by('name')
    )

    categories_list = []
    for category in categories_qs:
        related_fields = list(category.fields.all())
        related_field_count = len(related_fields)

        dynamic_schema = category.dynamic_fields or {}
        if isinstance(dynamic_schema, dict):
            json_field_count = len(dynamic_schema)
        elif isinstance(dynamic_schema, list):
            json_field_count = len(dynamic_schema)
        else:
            json_field_count = 0

        field_count = max(related_field_count, json_field_count)

        categories_list.append({
            'id': category.id,
            'name': category.name,
            'description': category.description or '',
            'asset_count': category.asset_count,
            'field_count': field_count,
        })
    
    # Get company context info
    company_info = {
        'name': company.name,
        'total_categories': len(categories_list),
        'total_assets': Asset.objects.for_company(company).count(),
        'total_fields': AssetCategoryField.objects.for_company(company).count(),
    }
    
    return JsonResponse({
        'success': True, 
        'categories': categories_list,
        'company': company_info
    })

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
@require_GET
def api_category_fields(request, category_id):
    """Enhanced API endpoint with field usage analytics."""
    company = getattr(request, 'company', None)
    if not company:
        return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
    try:
        category = AssetCategory.objects.for_company(company).get(pk=category_id)
    except AssetCategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Category not found.'}, status=404)
    
    fields = AssetCategoryField.objects.for_company(company).filter(category=category).order_by('label')
    
    # Calculate field usage statistics
    assets_in_category = Asset.objects.for_company(company).filter(category=category)
    total_assets = assets_in_category.count()
    
    data = []
    for f in fields:
        # Count how many assets have this field populated
        usage_count = 0
        if total_assets > 0:
            for asset in assets_in_category:
                if f.key in asset.dynamic_data and asset.dynamic_data[f.key]:
                    usage_count += 1
        
        usage_percentage = round((usage_count / total_assets * 100), 1) if total_assets > 0 else 0
        
        data.append({
            'id': f.id,
            'key': f.key,
            'label': f.label,
            'type': f.type,
            'required': f.required,
            'usage_count': usage_count,
            'usage_percentage': usage_percentage,
        })
    
    return JsonResponse({
        'success': True, 
        'fields': data,
        'category': {
            'id': category.id,
            'name': category.name,
            'total_assets': total_assets,
        }
    })

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
@require_POST
@csrf_protect
@transaction.atomic
def api_create_field(request, category_id):
    """
    Create a new field for a category.
    ENHANCED: Supports both JSON and form data for wizard compatibility.
    WORLD-CLASS: Comprehensive validation with detailed error messages.
    """
    import json
    import re
    from audit.utils import log_audit
    
    company = getattr(request, 'company', None)
    if not company:
        return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
    
    try:
        category = AssetCategory.objects.for_company(company).get(pk=category_id)
    except AssetCategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Category not found.'}, status=404)
    
    # Handle both JSON and form data (WORLD-CLASS FIX)
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            key = data.get('key', '').strip()
            label = data.get('label', '').strip()
            field_type = data.get('type', '').strip()
            required = data.get('required', False)
            if isinstance(required, str):
                required = required.lower() in ('true', '1', 'yes', 'on')
        else:
            key = request.POST.get('key', '').strip()
            label = request.POST.get('label', '').strip()
            field_type = request.POST.get('type', '').strip()
            required_raw = request.POST.get('required', 'false').lower()
            # Handle checkbox: 'on', 'true', '1', 'yes' = True, everything else = False
            required = required_raw in ('on', 'true', '1', 'yes')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data.'}, status=400)
    
    # WORLD-CLASS VALIDATION with specific error messages
    errors = []
    
    # Validate key
    if not key:
        errors.append('Field key is required.')
    elif len(key) < 2:
        errors.append('Field key must be at least 2 characters long.')
    elif len(key) > 50:
        errors.append('Field key must not exceed 50 characters.')
    elif not re.match(r'^[a-z][a-z0-9_]*$', key):
        errors.append('Field key must start with a lowercase letter and contain only lowercase letters, numbers, and underscores.')
    
    # Validate label
    if not label:
        errors.append('Field label is required.')
    elif len(label) < 2:
        errors.append('Field label must be at least 2 characters long.')
    elif len(label) > 100:
        errors.append('Field label must not exceed 100 characters.')
    
    # Validate field type
    valid_types = dict(AssetCategoryField.FIELD_TYPES)
    if not field_type:
        errors.append('Field type is required.')
    elif field_type not in valid_types:
        errors.append(f'Invalid field type. Must be one of: {", ".join(valid_types.keys())}.')
    
    # Check for duplicate key
    if key and AssetCategoryField.objects.for_company(company).filter(category=category, key__iexact=key).exists():
        errors.append(f'A field with key "{key}" already exists in this category.')
    
    # Return validation errors if any
    if errors:
        return JsonResponse({
            'success': False, 
            'error': ' '.join(errors),
            'errors': errors,
            'debug': {
                'key': key,
                'label': label,
                'type': field_type,
                'required': required,
                'valid_types': list(valid_types.keys())
            }
        }, status=400)
    
    # Create field
    try:
        field = AssetCategoryField.objects.create(
            company=company,
            category=category,
            key=key,
            label=label,
            type=field_type,
            required=required,
        )
        
        log_audit(request.user, 'create', None, f'Dynamic field created: {label} ({key}) in category {category.name}')
        
        return JsonResponse({
            'success': True,
            'field': {
                'id': field.id,
                'key': field.key,
                'label': field.label,
                'type': field.type,
                'required': field.required
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Failed to create field: {str(e)}'
        }, status=500)

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
@require_POST
@csrf_protect
@transaction.atomic
def api_update_field(request, field_id):
    """
    Update an existing category field.
    ENHANCED: Supports updating key for legacy fields, comprehensive validation.
    """
    import json
    import re
    from audit.utils import log_audit
    
    company = getattr(request, 'company', None)
    if not company:
        return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
    
    try:
        field = AssetCategoryField.objects.select_related('category').for_company(company).get(pk=field_id)
    except AssetCategoryField.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Field not found.'}, status=404)
    
    # Handle both JSON and form data
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            key = data.get('key', '').strip()
            label = data.get('label', '').strip()
            field_type = data.get('type', '').strip()
            required = data.get('required', False)
            if isinstance(required, str):
                required = required.lower() in ('true', '1', 'yes', 'on')
        else:
            key = request.POST.get('key', '').strip()
            label = request.POST.get('label', '').strip()
            field_type = request.POST.get('type', '').strip()
            required_raw = request.POST.get('required', 'false').lower()
            required = required_raw in ('on', 'true', '1', 'yes')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data.'}, status=400)
    
    # WORLD-CLASS VALIDATION with specific error messages
    errors = []
    
    # Validate key (allow updating for legacy fields)
    if key:  # Key is optional in update (can keep existing)
        if len(key) < 2:
            errors.append('Field key must be at least 2 characters long.')
        elif len(key) > 50:
            errors.append('Field key must not exceed 50 characters.')
        elif not re.match(r'^[a-z][a-z0-9_]*$', key):
            errors.append('Field key must start with a lowercase letter and contain only lowercase letters, numbers, and underscores.')
        # Check for duplicate key (excluding current field)
        elif AssetCategoryField.objects.for_company(company).filter(
            category=field.category, 
            key__iexact=key
        ).exclude(id=field_id).exists():
            errors.append(f'A field with key "{key}" already exists in this category.')
    
    # Validate label
    if not label:
        errors.append('Field label is required.')
    elif len(label) < 2:
        errors.append('Field label must be at least 2 characters long.')
    elif len(label) > 100:
        errors.append('Field label must not exceed 100 characters.')
    
    # Validate field type
    valid_types = dict(AssetCategoryField.FIELD_TYPES)
    if not field_type:
        errors.append('Field type is required.')
    elif field_type not in valid_types:
        errors.append(f'Invalid field type. Must be one of: {", ".join(valid_types.keys())}.')
    
    # Return validation errors if any
    if errors:
        return JsonResponse({
            'success': False,
            'error': ' '.join(errors),
            'errors': errors,
            'debug': {
                'field_id': field_id,
                'key': key,
                'label': label,
                'type': field_type,
                'required': required,
                'valid_types': list(valid_types.keys())
            }
        }, status=400)
    
    # Update field
    try:
        old_key = field.key
        old_label = field.label
        old_type = field.type
        old_required = field.required
        
        # Update key if provided (for legacy field migration)
        if key and key != field.key:
            field.key = key
        
        field.label = label
        field.type = field_type
        field.required = required
        field.save()
        
        # Build audit message
        changes = []
        if key and key != old_key:
            changes.append(f'key: {old_key} → {key}')
        if label != old_label:
            changes.append(f'label: {old_label} → {label}')
        if field_type != old_type:
            changes.append(f'type: {old_type} → {field_type}')
        if required != old_required:
            changes.append(f'required: {old_required} → {required}')
        
        change_summary = ', '.join(changes) if changes else 'no changes'
        
        log_audit(
            request.user, 
            'update', 
            None, 
            f'Dynamic field updated: {label} ({field.key}) in category {field.category.name}. Changes: {change_summary}'
        )
        
        return JsonResponse({
            'success': True,
            'field': {
                'id': field.id,
                'key': field.key,
                'label': field.label,
                'type': field.type,
                'required': field.required
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Failed to update field: {str(e)}'
        }, status=500)

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
@require_POST
@csrf_protect
@transaction.atomic
def api_delete_field(request, field_id):
    from audit.utils import log_audit
    company = getattr(request, 'company', None)
    if not company:
        return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
    try:
        field = AssetCategoryField.objects.select_related('category').for_company(company).get(pk=field_id)
    except AssetCategoryField.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Field not found.'}, status=404)
    category = field.category
    key = field.key
    label = field.label
    # Optionally: check if field is in use in any Asset.dynamic_data
    field.delete()
    log_audit(request.user, 'delete', None, f'Dynamic field deleted: {label} ({key}) in category {category.name}')
    return JsonResponse({'success': True})

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role in ('admin', 'manager'))
@require_POST
@csrf_protect
def asset_delete(request, asset_id):
    """
    DEPRECATED: Direct deletion is no longer recommended.
    
    World-Class Best Practice:
    - Use disposal workflow instead (soft delete with audit trail)
    - Preserves asset history and relationships
    - Requires admin approval
    - Maintains compliance
    
    This endpoint redirects to the disposal request workflow.
    For permanent deletion, super admins can use the dedicated endpoint.
    """
    try:
        asset = Asset.objects.filter(pk=asset_id).first()
        if not asset:
            return JsonResponse({'success': False, 'error': 'Asset not found.'}, status=404)
        
        # Log deprecated usage
        log_audit(
            request.user,
            'deprecated_delete_attempt',
            asset,
            f'User attempted deprecated direct delete. Redirected to disposal workflow.',
            company=asset.company,
            branch=asset.branch
        )
        
        # Return redirect to disposal workflow
        try:
            disposal_url = reverse('assets:asset_disposal_request', kwargs={'asset_uuid': asset.uuid})
        except NoReverseMatch:
            disposal_url = reverse('asset_disposal_request', kwargs={'asset_uuid': asset.uuid})

        # Allow disposal form to render even for admins when coming from the deprecated delete button.
        # This preserves a consistent UX while keeping the default admin behavior (edit/status change) elsewhere.
        separator = '&' if '?' in disposal_url else '?'
        disposal_url = f"{disposal_url}{separator}from=delete"

        return JsonResponse({
            'success': False,
            'deprecated': True,
            'message': 'Direct deletion is deprecated. Please use the disposal workflow to maintain audit trail and compliance.',
            'redirect_url': disposal_url,
            'action': 'redirect'
        }, status=410)  # 410 Gone - Resource no longer available
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role in ('admin', 'manager'))
@require_POST
@csrf_protect
@transaction.atomic
def asset_bulk_delete(request):
    """
    DEPRECATED: Bulk deletion is no longer recommended.
    Use bulk disposal workflow instead.
    """
    return JsonResponse({
        'success': False,
        'deprecated': True,
        'message': 'Bulk deletion is deprecated. Please use the disposal workflow for each asset to maintain audit trail.',
    }, status=410)


@login_required
@require_POST
@csrf_protect
def asset_permanent_delete(request, asset_id):
    """
    Permanently delete asset from database (RESTRICTED).
    
    Security & Compliance:
    - Super admin only
    - Asset must be archived for at least 90 days
    - Requires confirmation token
    - Full audit trail preserved
    
    Use Cases:
    - Test data cleanup
    - Duplicate removal
    - GDPR compliance (right to be forgotten)
    - Data retention policy enforcement
    
    URL: /assets/admin/permanent-delete/<id>/
    """
    # Super admin only
    if not request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'error': 'Permission denied. Only super admins can permanently delete assets.'
        }, status=403)
    
    try:
        company = getattr(request, 'company', None)
        asset = Asset.objects.get(pk=asset_id, company=company)
        
        # Must be archived (inactive)
        if asset.status not in ['retired', 'disposed', 'lost', 'archived']:
            return JsonResponse({
                'success': False,
                'error': 'Asset must be disposed/archived before permanent deletion.'
            }, status=400)
        
        # Check 90-day retention period
        if hasattr(asset, 'archived_at') and asset.archived_at:
            days_archived = (timezone.now() - asset.archived_at).days
            if days_archived < 90:
                return JsonResponse({
                    'success': False,
                    'error': f'Asset must be archived for at least 90 days. Currently: {days_archived} days.',
                    'days_remaining': 90 - days_archived
                }, status=400)
        
        # Require confirmation token
        confirmation = request.POST.get('confirmation', '')
        expected_token = f'DELETE-{asset.uuid}'
        if confirmation != expected_token:
            return JsonResponse({
                'success': False,
                'error': 'Invalid confirmation token.',
                'required_token': expected_token
            }, status=400)
        
        # Preserve asset data in audit log before deletion
        asset_data = {
            'id': asset.id,
            'uuid': str(asset.uuid),
            'description': str(asset),
            'status': asset.status,
            'company_id': asset.company_id,
            'branch_id': asset.branch_id if asset.branch else None,
            'category_id': asset.category_id if asset.category else None,
            'dynamic_data': asset.dynamic_data,
            'created_at': asset.created_at.isoformat() if asset.created_at else None,
        }
        
        # Log before deletion
        log_audit(
            request.user,
            'permanent_delete',
            asset,
            f'Asset permanently deleted from database: {asset}',
            company=company,
            branch=asset.branch,
            metadata={
                'asset_data': asset_data,
                'reason': request.POST.get('reason', 'Not specified'),
                'archived_at': asset.archived_at.isoformat() if hasattr(asset, 'archived_at') and asset.archived_at else None,
            }
        )
        
        asset_name = str(asset)
        asset.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Asset "{asset_name}" permanently deleted from database.',
            'warning': 'This action cannot be undone. Asset data has been preserved in audit log.'
        })
        
    except Asset.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Asset not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role in ('admin', 'manager'))
@require_POST
@csrf_protect
def regenerate_qr_code(request, uuid):
    """
    Regenerate QR code for a specific asset.
    This is useful when bulk import fails to generate QR codes or when QR codes need to be updated.
    """
    try:
        # Fetch asset by UUID with company scoping
        company = getattr(request, 'company', None)
        if not company:
            return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
        
        asset = Asset.objects.for_company(company).get(uuid=uuid)
        
        # Generate QR code with high quality settings
        try:
            import os
            from django.conf import settings
            
            # Ensure QR codes directory exists
            qr_dir = os.path.join(settings.MEDIA_ROOT, 'qr_codes')
            os.makedirs(qr_dir, exist_ok=True)
            
            # Build absolute URL for QR code
            base_url = request.build_absolute_uri('/')[:-1]
            qr_url = f"{base_url}/assets/{asset.uuid}/"
            
            # Generate QR code with high quality settings
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_H,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_url)
            qr.make(fit=True)
            
            img = qr.make_image(fill_color="black", back_color="white")
            buffer = BytesIO()
            img.save(buffer, 'PNG')
            buffer.seek(0)
            
            # Delete old QR code if exists
            if asset.qr_code:
                try:
                    asset.qr_code.delete(save=False)
                except Exception:
                    pass
            
            # Save new QR code to asset
            asset.qr_code.save(
                f"asset_{asset.uuid}.png",
                ContentFile(buffer.getvalue()),
                save=True
            )
            
            # Log the regeneration
            log_audit(
                request.user,
                'qr_regenerate',
                asset,
                f'QR code regenerated for asset {asset.pk}'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'QR code regenerated successfully.',
                'qr_code_url': asset.qr_code.url if asset.qr_code else None
            })
            
        except Exception as qr_error:
            return JsonResponse({
                'success': False,
                'error': f'QR code generation failed: {str(qr_error)}'
            }, status=500)
            
    except Asset.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Asset not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role in ('admin', 'manager'))
@require_POST
@csrf_protect
def bulk_regenerate_qr_codes(request):
    """
    Regenerate QR codes for multiple assets.
    Useful after bulk import or when QR codes need to be batch updated.
    """
    try:
        company = getattr(request, 'company', None)
        if not company:
            return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
        
        # Get asset IDs from request
        asset_ids = request.POST.getlist('asset_ids[]') or request.POST.getlist('asset_ids')
        if not asset_ids:
            return JsonResponse({'success': False, 'error': 'No asset IDs provided.'}, status=400)
        
        valid_ids = [int(i) for i in asset_ids if str(i).isdigit()]
        if not valid_ids:
            return JsonResponse({'success': False, 'error': 'No valid asset IDs provided.'}, status=400)
        
        # Fetch assets
        assets = Asset.objects.for_company(company).filter(pk__in=valid_ids)
        
        success_count = 0
        failed = []
        
        for asset in assets:
            try:
                import os
                from django.conf import settings
                
                # Ensure QR codes directory exists
                qr_dir = os.path.join(settings.MEDIA_ROOT, 'qr_codes')
                os.makedirs(qr_dir, exist_ok=True)
                
                # Build absolute URL for QR code
                base_url = request.build_absolute_uri('/')[:-1]
                qr_url = f"{base_url}/assets/{asset.uuid}/"
                
                # Generate QR code
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_H,
                    box_size=10,
                    border=4,
                )
                qr.add_data(qr_url)
                qr.make(fit=True)
                
                img = qr.make_image(fill_color="black", back_color="white")
                buffer = BytesIO()
                img.save(buffer, 'PNG')
                buffer.seek(0)
                
                # Delete old QR code if exists
                if asset.qr_code:
                    try:
                        asset.qr_code.delete(save=False)
                    except Exception:
                        pass
                
                # Save new QR code
                asset.qr_code.save(
                    f"asset_{asset.uuid}.png",
                    ContentFile(buffer.getvalue()),
                    save=True
                )
                
                success_count += 1
                
            except Exception as e:
                failed.append({'asset_id': asset.pk, 'error': str(e)})
        
        # Log bulk regeneration
        log_audit(
            request.user,
            'bulk_qr_regenerate',
            None,
            f'Bulk QR code regeneration: {success_count} successful, {len(failed)} failed',
            metadata={'success_count': success_count, 'failed_count': len(failed)}
        )
        
        return JsonResponse({
            'success': True,
            'message': f'QR codes regenerated for {success_count} asset(s).',
            'success_count': success_count,
            'failed': failed
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
@require_POST
@csrf_protect
def api_delete_category(request, category_id):
    """Delete a category (only if it has no assets)."""
    company = getattr(request, 'company', None)
    if not company:
        return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
    
    try:
        category = AssetCategory.objects.for_company(company).get(pk=category_id)
    except AssetCategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Category not found.'}, status=404)
    
    # Check if category has assets
    asset_count = Asset.objects.for_company(company).filter(category=category).count()
    if asset_count > 0:
        return JsonResponse({
            'success': False, 
            'error': f'Cannot delete category with {asset_count} asset(s). Please reassign or delete assets first.'
        }, status=400)
    
    category_name = category.name
    category.delete()
    log_audit(request.user, 'delete', None, f'Category deleted: {category_name} (ID: {category_id})')
    
    return JsonResponse({
        'success': True,
        'message': f'Category "{category_name}" deleted successfully.'
    })


@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
@require_GET
def api_category_analytics(request, category_id):
    """Get detailed analytics for a specific category."""
    company = getattr(request, 'company', None)
    if not company:
        return JsonResponse({'success': False, 'error': 'Company context required.'}, status=403)
    
    try:
        category = AssetCategory.objects.for_company(company).get(pk=category_id)
    except AssetCategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Category not found.'}, status=404)
    
    # Get assets in this category
    assets = Asset.objects.for_company(company).filter(category=category)
    total_assets = assets.count()
    
    # Status distribution
    status_distribution = assets.values('status').annotate(count=Count('id')).order_by('status')
    
    # Branch distribution
    branch_distribution = assets.values('branch__name').annotate(count=Count('id')).order_by('-count')[:5]
    
    # Recent activity (last 30 days)
    from datetime import timedelta
    from django.utils import timezone
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_assets = assets.filter(created_at__gte=thirty_days_ago).count()
    
    # Field usage statistics
    fields = AssetCategoryField.objects.for_company(company).filter(category=category)
    field_stats = []
    for field in fields:
        filled_count = sum(1 for asset in assets if field.key in asset.dynamic_data and asset.dynamic_data[field.key])
        usage_percentage = (filled_count / total_assets * 100) if total_assets > 0 else 0
        field_stats.append({
            'label': field.label,
            'key': field.key,
            'type': field.type,
            'required': field.required,
            'filled_count': filled_count,
            'usage_percentage': round(usage_percentage, 1)
        })
    
    # Get audit events for this category
    from audit.models import AuditLog
    recent_events = AuditLog.objects.filter(
        company=company,
        details__icontains=category.name
    ).order_by('-timestamp')[:10].values(
        'action', 'timestamp', 'user__username', 'details'
    )
    
    analytics = {
        'category': {
            'id': category.id,
            'name': category.name,
            'description': ''  # AssetCategory doesn't have description field
        },
        'summary': {
            'total_assets': total_assets,
            'total_fields': fields.count(),
            'recent_additions': recent_assets,
            'branches_using': branch_distribution.count()
        },
        'status_distribution': list(status_distribution),
        'branch_distribution': list(branch_distribution),
        'field_statistics': field_stats,
        'recent_activity': list(recent_events)
    }
    
    return JsonResponse({
        'success': True,
        'analytics': analytics
    })


@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
@require_GET
def api_category_templates(request):
    """Get available category templates for wizard."""
    from .category_templates import get_template_list
    
    templates = get_template_list()
    
    return JsonResponse({
        'success': True,
        'templates': templates
    })


@login_required
@user_passes_test(lambda u: u.is_authenticated and u.role == 'admin')
@require_GET
def api_category_template_detail(request, template_key):
    """Get detailed template configuration."""
    from .category_templates import get_template
    
    template = get_template(template_key)
    
    if not template:
        return JsonResponse({
            'success': False,
            'error': 'Template not found'
        }, status=404)
    
    return JsonResponse({
        'success': True,
        'template': template
    })


@login_required
@require_GET
def api_users_by_branch(request):
    """
    Get users filtered by branch for asset assignment.
    
    Query Params:
        branch_id: Branch ID to filter users
    
    Returns:
        JSON: List of users with id, username, role, branch info
    
    Security:
        - Company scoping enforced
        - Admins see all users (with cross-branch warnings)
        - Managers/Users see only users in selected branch
    """
    User = get_user_model()
    branch_id = request.GET.get('branch_id')
    company = getattr(request, 'company', None)
    user = request.user
    
    if not branch_id or not company:
        return JsonResponse({'users': []})
    
    try:
        branch_id = int(branch_id)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid branch_id'}, status=400)
    
    # Verify branch belongs to company (MULTI-TENANCY)
    from tenancy.models import Branch, UserBranch
    try:
        branch = Branch.objects.get(id=branch_id, company=company)
    except Branch.DoesNotExist:
        return JsonResponse({'error': 'Branch not found'}, status=404)
    
    # Admins see all users (with flag indicating if cross-branch)
    if user.role == 'admin':
        users_qs = User.objects.filter(
            company=company,
            is_active=True
        ).select_related('company').prefetch_related('user_branches__branch').order_by('username')
        
        users_data = []
        for u in users_qs:
            primary_branch = u.primary_branch
            is_in_selected_branch = False
            
            if primary_branch:
                is_in_selected_branch = primary_branch.id == branch_id
            
            users_data.append({
                'id': u.id,
                'username': u.username,
                'role': u.role,
                'branch_name': primary_branch.name if primary_branch else 'No Branch',
                'is_in_selected_branch': is_in_selected_branch,
                'display': f"{u.username} ({u.role}) - {primary_branch.name if primary_branch else 'No Branch'}"
            })
        
        return JsonResponse({'users': users_data, 'is_admin': True})
    
    # Managers/Users: Only users in selected branch (STRICT FILTERING)
    else:
        user_ids = UserBranch.objects.filter(
            branch_id=branch_id,
            company=company,
            branch__is_active=True
        ).values_list('user_id', flat=True)
        
        users_qs = User.objects.filter(
            id__in=user_ids,
            is_active=True
        ).select_related('company').prefetch_related('user_branches__branch').order_by('username')
        
        users_data = []
        for u in users_qs:
            primary_branch = u.primary_branch
            users_data.append({
                'id': u.id,
                'username': u.username,
                'role': u.role,
                'branch_name': primary_branch.name if primary_branch else 'No Branch',
                'is_in_selected_branch': True,
                'display': f"{u.username} ({u.role}) - {primary_branch.name if primary_branch else 'No Branch'}"
            })
        
        return JsonResponse({'users': users_data, 'is_admin': False})


@login_required
@company_required
def transfer_dashboard(request):
    """
    Transfer Management Dashboard - World-class approval interface.
    Shows pending approvals, transfer history, and comprehensive workflow management.
    """
    return render(request, 'assets/transfer_dashboard.html', {
        'page_title': 'Transfer Management',
    })


class AssetRegistrationWizardView(LoginRequiredMixin, TemplateView):
    """
    Phase 8: World-Class Asset Registration Wizard
    
    Multi-step wizard for asset registration with:
    - 6 steps: Basic Info, Location, Financial, Dynamic Fields, Documents, Review
    - Auto-save progress to localStorage
    - Real-time validation
    - Category-specific dynamic fields
    - File upload with preview
    - Review step before submission
    - Mobile-optimized
    - Accessibility (WCAG 2.1 AA)
    
    Inspired by: ServiceNow ITAM, IBM Maximo, SAP EAM
    """
    template_name = 'assets/asset_registration_wizard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = getattr(self.request, 'company', None)
        user = self.request.user
        
        # Get categories for the company
        context['categories'] = AssetCategory.objects.filter(
            company=company
        ).prefetch_related('fields').order_by('name')
        
        # Get branches for the company
        from tenancy.models import Branch
        context['branches'] = Branch.objects.filter(
            company=company,
            is_active=True
        ).order_by('name')
        
        # Get users for assignment (company-scoped)
        User = get_user_model()
        context['users'] = User.objects.filter(
            company=company,
            is_active=True
        ).order_by('username')
        
        # User role for frontend logic
        context['user_role'] = user.role
        
        return context
    
    def post(self, request, *args, **kwargs):
        """
        Handle wizard form submission
        
        Multi-tenancy: All data is company-scoped
        Security: CSRF protection, permission checks
        Validation: Server-side validation of all fields
        """
        company = getattr(request, 'company', None)
        user = request.user
        
        try:
            # Create asset using existing form logic
            form = AssetForm(request.POST, request.FILES)
            
            if form.is_valid():
                asset = form.save(commit=False)
                asset.company = company
                asset.created_by = user
                asset.updated_by = user
                
                # Generate QR code
                asset.save()
                
                # Generate QR code after save (needs ID)
                if not asset.qr_code:
                    qr = qrcode.QRCode(version=1, box_size=10, border=5)
                    qr.add_data(str(asset.uuid))
                    qr.make(fit=True)
                    img = qr.make_image(fill_color="black", back_color="white")
                    
                    buffer = BytesIO()
                    img.save(buffer, format='PNG')
                    file_name = f'qr_codes/{asset.uuid}.png'
                    asset.qr_code.save(file_name, ContentFile(buffer.getvalue()), save=True)
                
                # Log audit
                log_audit(
                    user, 
                    'create', 
                    asset,
                    f"Asset created via wizard: {asset.name}",
                    company=company,
                    branch=asset.branch,
                    metadata={
                        'asset_uuid': str(asset.uuid),
                        'category': asset.category.name,
                        'status': asset.status,
                        'method': 'wizard'
                    }
                )
                
                messages.success(
                    request, 
                    f'Asset "{asset.name}" registered successfully! QR code generated.'
                )
                return redirect('asset_detail_by_uuid', uuid=asset.uuid)
            else:
                # Form validation errors
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
                return self.render_to_response(self.get_context_data())
                
        except Exception as e:
            messages.error(request, f'Error registering asset: {str(e)}')
            return self.render_to_response(self.get_context_data())
